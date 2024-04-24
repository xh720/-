# -*- coding: utf-8 -*-
import sys
import os
from PyQt5.QtSerialPort import *
import xlrd
import openpyxl
from decimal import Decimal
from PyQt5.QtWidgets import QVBoxLayout,QGroupBox, QComboBox
import modbus_tk
import modbus_tk.defines as cst
from modbus_tk import modbus_rtu
import time
from SerialPortDialogFunc import SerialPortDialogFunc
from NewProgramDialogFunc import NewProgramDialogFunc

from Delete_Program_Func import Delete_ProgramDialogFunc

from AboutFunc import AboutFunc
import configparser
import serial  # 导入串口模块
import serial.tools.list_ports
from tkinter import filedialog
from PyQt5 import QtWidgets,QtCore
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
import 界面备份9

# ---------------------------新添加的东西---------------------------
aucCRCHi = [
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x00, 0xC1, 0x81, 0x40,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x00, 0xC1, 0x81, 0x40,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x00, 0xC1, 0x81, 0x40,
    0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41, 0x00, 0xC1, 0x81, 0x40,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x00, 0xC1, 0x81, 0x40,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41,
    0x01, 0xC0, 0x80, 0x41, 0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x00, 0xC1, 0x81, 0x40,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40, 0x01, 0xC0, 0x80, 0x41, 0x01, 0xC0, 0x80, 0x41,
    0x00, 0xC1, 0x81, 0x40]
aucCRCLo = [
    0x00, 0xC0, 0xC1, 0x01, 0xC3, 0x03, 0x02, 0xC2, 0xC6, 0x06, 0x07, 0xC7,
    0x05, 0xC5, 0xC4, 0x04, 0xCC, 0x0C, 0x0D, 0xCD, 0x0F, 0xCF, 0xCE, 0x0E,
    0x0A, 0xCA, 0xCB, 0x0B, 0xC9, 0x09, 0x08, 0xC8, 0xD8, 0x18, 0x19, 0xD9,
    0x1B, 0xDB, 0xDA, 0x1A, 0x1E, 0xDE, 0xDF, 0x1F, 0xDD, 0x1D, 0x1C, 0xDC,
    0x14, 0xD4, 0xD5, 0x15, 0xD7, 0x17, 0x16, 0xD6, 0xD2, 0x12, 0x13, 0xD3,
    0x11, 0xD1, 0xD0, 0x10, 0xF0, 0x30, 0x31, 0xF1, 0x33, 0xF3, 0xF2, 0x32,
    0x36, 0xF6, 0xF7, 0x37, 0xF5, 0x35, 0x34, 0xF4, 0x3C, 0xFC, 0xFD, 0x3D,
    0xFF, 0x3F, 0x3E, 0xFE, 0xFA, 0x3A, 0x3B, 0xFB, 0x39, 0xF9, 0xF8, 0x38,
    0x28, 0xE8, 0xE9, 0x29, 0xEB, 0x2B, 0x2A, 0xEA, 0xEE, 0x2E, 0x2F, 0xEF,
    0x2D, 0xED, 0xEC, 0x2C, 0xE4, 0x24, 0x25, 0xE5, 0x27, 0xE7, 0xE6, 0x26,
    0x22, 0xE2, 0xE3, 0x23, 0xE1, 0x21, 0x20, 0xE0, 0xA0, 0x60, 0x61, 0xA1,
    0x63, 0xA3, 0xA2, 0x62, 0x66, 0xA6, 0xA7, 0x67, 0xA5, 0x65, 0x64, 0xA4,
    0x6C, 0xAC, 0xAD, 0x6D, 0xAF, 0x6F, 0x6E, 0xAE, 0xAA, 0x6A, 0x6B, 0xAB,
    0x69, 0xA9, 0xA8, 0x68, 0x78, 0xB8, 0xB9, 0x79, 0xBB, 0x7B, 0x7A, 0xBA,
    0xBE, 0x7E, 0x7F, 0xBF, 0x7D, 0xBD, 0xBC, 0x7C, 0xB4, 0x74, 0x75, 0xB5,
    0x77, 0xB7, 0xB6, 0x76, 0x72, 0xB2, 0xB3, 0x73, 0xB1, 0x71, 0x70, 0xB0,
    0x50, 0x90, 0x91, 0x51, 0x93, 0x53, 0x52, 0x92, 0x96, 0x56, 0x57, 0x97,
    0x55, 0x95, 0x94, 0x54, 0x9C, 0x5C, 0x5D, 0x9D, 0x5F, 0x9F, 0x9E, 0x5E,
    0x5A, 0x9A, 0x9B, 0x5B, 0x99, 0x59, 0x58, 0x98, 0x88, 0x48, 0x49, 0x89,
    0x4B, 0x8B, 0x8A, 0x4A, 0x4E, 0x8E, 0x8F, 0x4F, 0x8D, 0x4D, 0x4C, 0x8C,
    0x44, 0x84, 0x85, 0x45, 0x87, 0x47, 0x46, 0x86, 0x82, 0x42, 0x43, 0x83,
    0x41, 0x81, 0x80, 0x40]


# modbus crc计算函数
def modbus_crc(data):
    """定义了一个名为modbus_crc的函数，
    用于计算Modbus通信协议中的
    CRC（循环冗余校验）值。"""
    CRCHi = 0xFF
    CRCLo = 0xFF
    index = 0
    # 使用异或(^ )操作和预定义的查找表(aucCRCHi和aucCRCLo)来计算新的CRC值。
    for i in range(0, len(data)):
        index = CRCLo ^ data[i]
        CRCLo = CRCHi ^ aucCRCHi[index]
        CRCHi = aucCRCLo[index]

    return CRCHi, CRCLo

class Pyqt5_Serial(QtWidgets.QWidget, 界面备份9.Ui_Form):
    def __init__(self):
        super(Pyqt5_Serial, self).__init__()
        # 临时计数
        self._send_id = 0
        self._translate = QtCore.QCoreApplication.translate
        self.ser = None
        comboBoxList1 = [self._translate("Form", "系统操作"), self._translate("Form", "流程控制"),
                         self._translate("Form", "输出口设置"), self._translate("Form", "回零运动"),
                         self._translate("Form", "插补运动"), self._translate("Form", "独立运动")]
        comboBoxList2 = {self._translate("Form", "系统操作"): [self._translate("Form", "停止"), self._translate("Form", "启动"),
                                                           self._translate("Form", "暂停"), self._translate("Form", "恢复"),
                                                           self._translate("Form", "延时等待"),
                                                           self._translate("Form", "等待电机完成"),
                                                           self._translate("Form", "停止电机运动"),
                                                           self._translate("Form", "常等待")],
                         self._translate("Form", "流程控制"): [self._translate("Form", "程序间跳转"),
                                                           self._translate("Form", "程序循环"),
                                                           self._translate("Form", "输入跳转"),
                                                           self._translate("Form", "开启输入中断"),
                                                           self._translate("Form", "关闭输入中断")],
                         self._translate("Form", "输出口设置"): [self._translate("Form", "输出口设置")],
                         self._translate("Form", "回零运动"): [self._translate("Form", "设置回零速度"),
                                                           self._translate("Form", "启动回零")],
                         self._translate("Form", "插补运动"): [self._translate("Form", "设置点位速度"),
                                                           self._translate("Form", "三轴相对运动"),
                                                           self._translate("Form", "单轴绝对运动"),
                                                           self._translate("Form", "XY绝对运动"),
                                                           self._translate("Form", "XZ绝对运动"),
                                                           self._translate("Form", "YZ绝对运动"),
                                                           self._translate("Form", "三轴绝对运动"),
                                                           self._translate("Form", "XY圆弧插补"),
                                                           self._translate("Form", "XZ圆弧插补"),
                                                           self._translate("Form", "YZ圆弧插补")],
                         self._translate("Form", "独立运动"): [self._translate("Form", "独立运动速度"),
                                                           self._translate("Form", "相对运动"),
                                                           self._translate("Form", "X绝对运动"),
                                                           self._translate("Form", "Y绝对运动"),
                                                           self._translate("Form", "Z绝对运动")]}

        self.setupUi(self)
        self.TableWidget.itemChanged.connect(self.on_TableWidgetX_itemChanged)
        self.port = ""
        self.readini = configparser.ConfigParser()
        self.conf_path = os.path.join(os.path.abspath(os.path.dirname(sys.argv[0])), "APP.ini")
        self.ProjectName = ""
        self.Language = "CHINESE"
        self.readini.read(self.conf_path)
        #读取上次内容
        self.port = self.readini.get("common","port")
        self.botelv = self.readini.get("common","botelv")
        self.shujuwei = self.readini.get("common","shujuwei")
        self.jiaoyanwei = self.readini.get("common", "jiaoyanwei")
        self.stop = self.readini.get("common", "stop")

        print(self.port)
        les = self.findChildren(QLineEdit)
        for item in les:
            if item.property("type") != None:
                item.installEventFilter(self)

        self.init()
        self.setWindowTitle("三轴电机控制软件")
        # self.setWindowIcon(QIcon('D:/个人文件/软件图标.png'))

        self.X_minus_button.setProperty("name","X")
        self.X_minus_button.setProperty("method","sub")
        self.X_plus_button.setProperty("name","X")
        self.X_plus_button.setProperty("method","plus")

        self.Y_minus_button.setProperty("name", "Y")
        self.Y_minus_button.setProperty("method", "sub")
        self.Y_plus_button.setProperty("name", "Y")
        self.Y_plus_button.setProperty("method", "plus")

        self.Z_minus_button.setProperty("name", "Z")
        self.Z_minus_button.setProperty("method", "sub")
        self.Z_plus_button.setProperty("name", "Z")
        self.Z_plus_button.setProperty("method", "plus")

        self.cupboard = []

        self.timer1 = 0
        self.timer2 = 0
    def waitTime(self,waittime:int = 2000):
        loop = QEventLoop()  # 时间的卡顿
        QTimer.singleShot(waittime,loop.quit)  #卡子的退出
        loop.exec()  #  卡子

    def on_TableWidgetX_itemChanged(self,item):
        """用于处理TableWidget（可能是Qt框架中的一个表格控件）中的单元格内容变化事件。
        当表格中某个单元格的内容发生变化时，这个方法会被调用。"""
        # 获取发生变化的单元格项item所在的行号，并将其存储再row变量中
        row = self.TableWidget.row(item)
        # 获取了与发生变化的单元格项item位于同一行的第0列和第1列的单元格项
        item1 = self.TableWidget.item(row, 0)
        item2 = self.TableWidget.item(row, 1)
        if item1 != None and item2 != None:
            # 依次调用handlecolumn3到handlecolumn7这五个方法，
            self.handlecolumn3(item)
            self.handlecolumn4(item)
            self.handlecolumn5(item)
            self.handlecolumn6(item)
            self.handlecolumn7(item)

    def handlecolumn3(self, item):
        """
        参数1
        :param item:
        :return:
        """
        # 处理特定列中单元格内容变化情况
        if self.TableWidget.column(item) == 2:
            # 将获取该单元格的文本内容，并将其存储在变量value中
            value = item.text()
            # 获取该单元格项item所在的行号，并将其存储在变量row中
            row = self.TableWidget.row(item)
            # 获取与单元格项item同一行的第1列的单元格项，并获取该单元格项的文本内容，将其存储在变量firstItem中
            firstItem = self.TableWidget.item(row, 0).text()
            secondItem = self.TableWidget.item(row, 1).text()

            """ 检查与发生变化的单元格项item同一行的第1列和第2列单元格的文本内容是否分别等于特定的字符串。
                        通过self._translate方法获取，将字符串从原始语言翻译成当前界面使用的语言"""

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止"):
                # 将第3列单元格项item的文本内容（存储在变量value中）转换成一个整数，并将结果存储在变量Intvalue中
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "启动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "暂停"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "恢复"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "延时等待"):
                Floatvalue = float(value)
                if Floatvalue < 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "等待电机完成"):
                Intvalue = int(value)
                if Intvalue != 0 and Intvalue != 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止电机运动"):
                Intvalue = int(value)
                if Intvalue != 0 and Intvalue != 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "常等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序间跳转"):
                Intvalue = int(value)
                if Intvalue < 1 or Intvalue > 255:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序循环"):
                Intvalue = int(value)
                if Intvalue < 1 or Intvalue > 255:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "输入跳转"):
                Intvalue = int(value)
                if Intvalue < 1 or Intvalue > 4:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
                    # self.TableWidget.openPersistentEditor(item)

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "开启输入中断"):
                Intvalue = int(value)
                if Intvalue < 1 or Intvalue > 4:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
                    # self.TableWidget.openPersistentEditor(item)

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "关闭输入中断"):
                Intvalue = int(value)
                if Intvalue < 1 or Intvalue > 4:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
                    # self.TableWidget.openPersistentEditor(item)
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "输出口设置") and secondItem == self._translate("Form", "输出口设置"):
                Intvalue = int(value)
                if Intvalue < 1 or Intvalue > 4:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "设置回零速度"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "启动回零"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "设置点位速度"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴相对运动"):
                print("都可以")
                # Intvalue = int(value)
                # if Intvalue <= 0:
                #     QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "单轴绝对运动"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 3:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "独立运动速度"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 3:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "相对运动"):
                print("随便！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "X绝对运动"):
                Floatvalue = float(value)
                if Floatvalue < 0.0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Y绝对运动"):
                Floatvalue = float(value)
                if Floatvalue < 0.0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Z绝对运动"):
                Floatvalue = float(value)
                if Floatvalue < 0.0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

    def handlecolumn4(self, item):
        """
        参数2
        :param item:
        :return:
        """
        if self.TableWidget.column(item) == 3:
            value = item.text()
            row = self.TableWidget.row(item)
            firstItem = self.TableWidget.item(row, 0).text()
            secondItem = self.TableWidget.item(row, 1).text()
            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "启动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "暂停"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "恢复"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "延时等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "等待电机完成"):
                Intvalue = int(value)
                if Intvalue != 0 and Intvalue != 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止电机运动"):
                Intvalue = int(value)
                if Intvalue != 0 and Intvalue != 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "常等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序间跳转"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误!")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序循环"):
                Intvalue = int(value)
                if Intvalue < 1 or Intvalue > 255:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "输入跳转"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误!")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "开启输入中断"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误!")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "关闭输入中断"):
                Intvalue = int(value)
                if Intvalue != 0 :
                    QMessageBox.critical(self, "Port Error\n", "数据有误!")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "输出口设置") and secondItem == self._translate("Form", "输出口设置"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
                    # self.TableWidget.openPersistentEditor(item)
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "设置回零速度"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
                    # self.TableWidget.openPersistentEditor(item)

            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "启动回零"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "设置点位速度"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴相对运动"):
                print("都可以")
                # Intvalue = int(value)
                # if Intvalue <= 0:
                #     QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "单轴绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "独立运动速度"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "相对运动"):
                print("随便吧！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "X绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Y绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Z绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

    def handlecolumn5(self, item):
        """
        参数3
        :param item:
        :return:
        """
        if self.TableWidget.column(item) == 4:
            value = item.text()
            row = self.TableWidget.row(item)
            firstItem = self.TableWidget.item(row, 0).text()
            secondItem = self.TableWidget.item(row, 1).text()
            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "启动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "暂停"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "恢复"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "延时等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "等待电机完成"):
                Intvalue = int(value)
                if Intvalue != 0 and Intvalue != 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止电机运动"):
                Intvalue = int(value)
                if Intvalue != 0 and Intvalue != 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "常等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序间跳转"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误!")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序循环"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "输入跳转"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 255:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "开启输入中断"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 255:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "关闭输入中断"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "输出口设置") and secondItem == self._translate("Form", "输出口设置"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "设置回零速度"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
                    # self.TableWidget.openPersistentEditor(item)

            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "启动回零"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "设置点位速度"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴相对运动"):
                print("都可以")
                # Intvalue = int(value)
                # if Intvalue <= 0:
                #     QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "单轴绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴绝对运动"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "独立运动速度"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "相对运动"):
                print("随便吧！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "X绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0 :
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Y绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Z绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

    def handlecolumn6(self, item):
        """
        备用1
        :param item:
        :return:
        """
        if self.TableWidget.column(item) == 5:
            value = item.text()
            row = self.TableWidget.row(item)
            firstItem = self.TableWidget.item(row, 0).text()
            secondItem = self.TableWidget.item(row, 1).text()
            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "启动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "暂停"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "恢复"):
                Intvalue = int(value)
                if Intvalue != 0 :
                    QMessageBox.critical(self, "Port Error\n", "数据有错！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "延时等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "等待电机完成"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止电机运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "常等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序间跳转"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误!")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序循环"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "输入跳转"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "开启输入中断"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "关闭输入中断"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "输出口设置") and secondItem == self._translate("Form", "输出口设置"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "设置回零速度"):
                Floatvalue = float(value)
                if Floatvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "启动回零"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "设置点位速度"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴相对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "单轴绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ圆弧插补"):
                Floatvalue = float(value)
                if Floatvalue <= 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "独立运动速度"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "相对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "X绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Y绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Z绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

    def handlecolumn7(self, item):
        """
        备用2
        :param item:
        :return:
        """
        if self.TableWidget.column(item) == 6:
            value = item.text()
            row = self.TableWidget.row(item)
            firstItem = self.TableWidget.item(row, 0).text()
            secondItem = self.TableWidget.item(row, 1).text()
            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "启动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "暂停"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "恢复"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "延时等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "等待电机完成"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "停止电机运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "系统操作") and secondItem == self._translate("Form", "常等待"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序间跳转"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误!")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "程序循环"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "输入跳转"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "开启输入中断"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")

            if firstItem == self._translate("Form", "流程控制") and secondItem == self._translate("Form", "关闭输入中断"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "有错！")
            # ------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "输出口设置") and secondItem == self._translate("Form", "输出口设置"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # -----------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "设置回零速度"):
                Floatvalue = float(value)
                if Floatvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "回零运动") and secondItem == self._translate("Form", "启动回零"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # -------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "设置点位速度"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴相对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "单轴绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "三轴绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XY圆弧插补"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "XZ圆弧插补"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "插补运动") and secondItem == self._translate("Form", "YZ圆弧插补"):
                Intvalue = int(value)
                if Intvalue < 0 or Intvalue > 1:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")
            # --------------------------------------------------------------------------------------
            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "独立运动速度"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "相对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "X绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Y绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

            if firstItem == self._translate("Form", "独立运动") and secondItem == self._translate("Form", "Z绝对运动"):
                Intvalue = int(value)
                if Intvalue != 0:
                    QMessageBox.critical(self, "Port Error\n", "数据有误！")

    # 用于处理定时器事件
    # 当定时器事件发生时，timerEvent方法被调用，并传入一个QTimerEvent对象a0
    def timerEvent(self, a0: 'QTimerEvent'):
        # 检查触发此事件的定时器是否是timer1
        if a0.timerId() == self.timer1:
            # 如果是timer1，调用refresh_manual_data方法
            self.refresh_manual_data()

        # 检查触发此事件的定时器是否是timer2
        if a0.timerId() == self.timer2:
            # 如果是timer2，调用Line_Arc_Para_Read方法
            self.Line_Arc_Para_Read()

        # if a0.timerId() == self.timer3:
        #     self.Editing_program_show_data()

    def init(self):
        # 将open_button按钮的clicked信号连接到port_open槽函数
        self.open_button.clicked.connect(self.port_open)
        # 创建了一个QTranslator对象并将其赋值给self.trans
        self.trans = QTranslator()
        # 将BackZero_button按钮的clicked信号连接到STOP槽函数
        self.BackZero_button.clicked.connect(self.STOP)
        self.Download_program_button.clicked.connect(self.data_send)
        # 将Enter4_button按钮的clicked信号连接到Line_Arc_data_send槽函数
        self.Enter4_button.clicked.connect(self.Line_Arc_data_send)
        ########################################################################
        # 定时发送数据
        self.timer_send = QTimer()
        # todo 参数设置按钮绑定function
        self.Read_button.clicked.connect(self.Line_Arc_Para_Read)
        self.Line_Arc_button.clicked.connect(self.Line_Arc_interpolation)
        # self.TableWidget.itemDoubleClicked.connect(self.doubleclicked)  #双击TableWidget 暂停刷新
        self.Peizhi_button.clicked.connect(self.peizhichuankou)  #串口配置按键的连接函数
        self.stop_button.clicked.connect(self.port_close)
        self.About_button.clicked.connect(self.about)
        self.Manual_input_button.clicked.connect(self.Manual_input_show)
        self.Parameter_setting_button.clicked.connect(self.Parameter_setting_show)
        self.Editing_program_button.clicked.connect(self.Editing_program_show)
        self.Editing_program_button.clicked.connect(self.on_btnAutoHeight_clicked)
        self.Line_up_button.clicked.connect(self.moveup)
        self.Line_down_button.clicked.connect(self.movedown)
        self.Come_in_button.clicked.connect(self.AddRow)
        self.Delete_1_button.clicked.connect(self.DeleteRow)
        self.Copy_button.clicked.connect(self.CopyRow)
        self.Paste_button.clicked.connect(self.PasteRow)
        self.Save_program_button.clicked.connect(self.saveExcelfiles)
        self.Open_program_button.clicked.connect(self.openExcelfiles)
        self.Enter1_button.clicked.connect(self.Enter_speed)
        self.Enter2_button.clicked.connect(self.Enter_bujin)
        self.Para_read_button.clicked.connect(self.Parameter_Read)
        self.Para_write_button.clicked.connect(self.Parameter_Write)

        # 将self.X_minus_button按钮的downpressed信号连接到self.minus_down槽函数
        self.X_minus_button.downpressed.connect(self.minus_down)
        self.X_minus_button.uppressed.connect(self.plus_up)
        self.X_plus_button.downpressed.connect(self.minus_down)
        self.X_plus_button.uppressed.connect(self.plus_up)
        self.Y_minus_button.downpressed.connect(self.minus_down)
        self.Y_minus_button.uppressed.connect(self.plus_up)
        self.Y_plus_button.downpressed.connect(self.minus_down)
        self.Y_plus_button.uppressed.connect(self.plus_up)
        self.Z_minus_button.downpressed.connect(self.minus_down)
        self.Z_minus_button.uppressed.connect(self.plus_up)
        self.Z_plus_button.downpressed.connect(self.minus_down)
        self.Z_plus_button.uppressed.connect(self.plus_up)

        self.BackX_button.clicked.connect(self.Back_X_zero)
        self.BackY_button.clicked.connect(self.Back_Y_zero)
        self.BackZ_button.clicked.connect(self.Back_Z_zero)
        self.New_program_button.clicked.connect(self.New_program)
        # self.New_program_button.clicked.connect(self.on_btnAutoHeight_clicked)
        self.Delete_program_button.clicked.connect(self.Delete_program)
        # 将CheckBox_1复选框的stateChanged信号连接到set_qcheck_state1槽函数。
        # 当CheckBox_1的状态改变时（比如从选中变为未选中，或反之），set_qcheck_state1函数将被调用。
        self.CheckBox_1.stateChanged.connect(self.set_qcheck_state1)
        self.CheckBox_2.stateChanged.connect(self.set_qcheck_state2)
        self.CheckBox_3.stateChanged.connect(self.set_qcheck_state3)
        self.CheckBox_4.stateChanged.connect(self.set_qcheck_state4)
        self.CheckBox_19.stateChanged.connect(self.set_qcheck_state5)
        self.CheckBox_20.stateChanged.connect(self.set_qcheck_state6)
        # verticalGroupBox_3这个组框变为可见。
        # setVisible(True)表示设置该组框为可见状态。
        self.verticalGroupBox_3.setVisible(True)
        self.GroupBoxPara.setVisible(False)
        self.GroupBoxEditProgram.setVisible(False)
        self.GroupBoxLine_arc.setVisible(False)
        # 将六个文本输入框（lineEdit_13、lineEdit_16等）的textEdited信号连接到同一个槽函数self.onChange。
        # 使用了lambda表达式来简化连接过程
        self.lineEdit_13.textEdited.connect(lambda: self.onChange())
        self.lineEdit_16.textEdited.connect(lambda: self.onChange())
        self.lineEdit_14.textEdited.connect(lambda: self.onChange())
        self.lineEdit_17.textEdited.connect(lambda: self.onChange())
        self.lineEdit_15.textEdited.connect(lambda: self.onChange())
        self.lineEdit_18.textEdited.connect(lambda: self.onChange())
        # 将s1__box_23（可能是一个下拉框或组合框）的currentIndexChanged信号连接到self.selectionchange槽函数。
        self.s1__box_23.currentIndexChanged.connect(self.selectionchange)
        self.Para_save_button1.clicked.connect(self.Para_save)
        self.Para_save_button2.clicked.connect(self.Para_save)
        self.Para_save_button4.clicked.connect(self.Para_save)
        #新界面的信号槽
        self.XY_Line_interpolation_button.clicked.connect(self.XY_Line_interpolation)
        self.XZ_Line_interpolation_button.clicked.connect(self.XZ_Line_interpolation)
        self.YZ_Line_interpolation_button.clicked.connect(self.YZ_Line_interpolation)
        self.XY_Arc_interpolation_button.clicked.connect(self.XY_Arc_interpolation)
        self.XZ_Arc_interpolation_button.clicked.connect(self.XZ_Arc_interpolation)
        self.YZ_Arc_interpolation_button.clicked.connect(self.YZ_Arc_interpolation)
        self.Fast_Stop_button.clicked.connect(self.STOP)
        self.Return_main_interface_button.clicked.connect(self.Main_interface)
        self.Basic_Para_button.clicked.connect(self.Basic_para_interface)

    ###############################################################
    def Line_Arc_data_send(self):
        """使用 modbus_tk.utils.create_logger函数创建了一个日志记录器，并将其命名为 "console"。
        接着，它记录了一条信息级别的日志，内容为 "connected"，表示已经连接。"""
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        # 从lineEdit_34的文本输入框中获取文本内容，并将其转换为浮点数，然后赋值给变量 a_1
        a_1 = float(self.lineEdit_34.text())
        b_1 = a_1 * 1000
        # 计算b_1除以65536的余数，并将结果转换为整数，赋值给 c_1。
        # 65536是2的16次方，所以这种取余操作可能是为了得到一个 16 位的数据值。
        c_1 = int(b_1 % 65536)
        print(c_1)
        # 执行 b_1除以 65536 的整数除法，并将结果转换为整数，赋值给 d_1。
        # 整数除法会返回商的整数部分，因此这里可能是为了得到一个高位（超过 16 位）的数据值
        d_1 = int(b_1 // 65536)
        print(d_1)
        """for i in range(1): 是一个只循环一次的循环，因为 range(1) 只生成一个数字 0，这通常用于确保某段代码只执行一次。
logger.info(...) 记录日志信息。它记录 master.execute(...) 方法的返回值，该方法执行 Modbus 写入操作。
master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2179, output_value=[c_1, d_1]) 调用了 execute 方法来执行一个 Modbus 操作。"""
        # 1 可能是从站地址（slave address）。
        # cst.WRITE_MULTIPLE_REGISTERS 是功能码，表示要执行的操作是写入多个寄存器。
        # 0x2179 是起始寄存器的地址。
        # output_value=[c_1, d_1] 是要写入寄存器的值，这里是一个包含两个元素的列表，分别是 c_1 和 d_1
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2179, output_value=[c_1, d_1]))

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        a_2 = float(self.lineEdit_35.text())
        b_2 = a_2 * 1000
        c_2 = int(b_2 % 65536)
        print(c_2)
        d_2 = int(b_2 // 65536)
        print(d_2)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x21A2, output_value=[c_2, d_2]))

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        a_3 = float(self.lineEdit_36.text())
        b_3 = a_3 * 1000
        c_3 = int(b_3 % 65536)
        print(c_3)
        d_3 = int(b_3 // 65536)
        print(d_3)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x21A4, output_value=[c_3, d_3]))

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        a_4 = float(self.lineEdit_37.text())
        b_4 = a_4 * 1000
        c_4 = int(b_4 % 65536)
        print(c_4)
        d_4 = int(b_4 // 65536)
        print(d_4)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x21A6, output_value=[c_4, d_4]))

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        a_5 = float(self.lineEdit_38.text())
        b_5 = a_5 * 1000
        c_5 = int(b_5 % 65536)
        print(c_5)
        d_5 = int(b_5 // 65536)
        print(d_5)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x21A8, output_value=[c_5, d_5]))

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        a_6 = float(self.lineEdit_39.text())
        b_6 = a_6 * 1000
        c_6 = int(b_6 % 65536)
        print(c_6)
        d_6 = int(b_6 // 65536)
        print(d_6)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x21AA, output_value=[c_6, d_6]))

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        a_7 = float(self.lineEdit_40.text())
        b_7 = a_7 * 1000
        c_7 = int(b_7 % 65536)
        print(c_7)
        d_7 = int(b_7 // 65536)
        print(d_7)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x21AC, output_value=[c_7, d_7]))

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        a_8 = int(self.lineEdit_44.text())
        # 如果 a_8 等于 0，则执行 Modbus 单个寄存器写入操作，将值 0 写入地址为 0x21AE 的寄存器
        if a_8 == 0:
            logger.info(master.execute(1, cst.WRITE_SINGLE_REGISTER, 0x21AE, output_value=0))
        if a_8 == 1:
            logger.info(master.execute(1, cst.WRITE_SINGLE_REGISTER, 0x21AE, output_value=1000))

        # 调用startTimer方法启动一个定时器，定时器的间隔为100毫秒，并将返回的定时器ID赋值给self.timer2
        if self.timer2==0:
            self.timer2 = self.startTimer(100)

    def Line_Arc_Para_Read(self):
        # todo 在这个位置修改程序，读取数据时只读取手动操作界面的数据
        if self.timer2 == 0:
            self.timer2 = self.startTimer(100)
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器

        # todo 因为要将从Modbus Slave 读取到的数据放到手动操作的各个文本框中，而现在读取到的数据在表格中，并且分割了五份，
        # todo 要将读取到的数据分别放到5个文本框和勾选输入口状态和输出口状态
        # todo 在Modbus Slave中写寄存器数据的时候
        # todo 前0个寄存器是对应的“上电自动运行”
        # todo 第1个寄存器到第10个寄存器数据对应的是 5个文本框
        # todo 第11个寄存器到第14个寄存器数据对应的是 “输出口选择”
        # todo 第15个寄存器到第21个寄存器数据对应的是 “输入口状态”
        # todo 第22个寄存器到第25个寄存器数据对应的是 “输出口状态”
        # todo 第26个寄存器到第28个寄存器数据对应的是 “电机状态”
        # 读取 Modbus 服务器上的寄存器值，并将读取到的数据转换为字符串列表
        # 1从站地址  cst.HOLDING_REGISTERS功能码  0x21A2起始寄存器地址  13读取寄存器数量
        read = master.execute(1, cst.HOLDING_REGISTERS, 0x21A2, 13)
        print(read)
        strRead = [str(i) for i in read[0]]
        print(strRead)

        """int(strRead[1]) * 65536 将第二个寄存器值（通常是一个16位值的高8位）乘以65536（即2的16次方），这是为了将其转换回其原始的高位值。
int(strRead[1]) * 65536 + int(strRead[0]) 将转换后的高位值和低位值（strRead[0]）相加，从而得到一个完整的32位（或更少的位数，取决于实际使用的位数）值。"""
        #X轴的终点坐标
        info1 = (int(strRead[1]) * 65536 + int(strRead[0])) / 1000
        # %.3f 是一个格式化字符串，用于将浮点数格式化为具有三位小数的字符串。
        # "%.3f" % info1 将 info1 的值格式化为一个字符串，该字符串具有三位小数。
        # self.lineEdit_35.setText(...) 将格式化后的字符串设置为 lineEdit_35 文本框的内容。
        self.lineEdit_35.setText("%.3f" % info1)

        # Y轴的终点坐标
        info2 = (int(strRead[3]) * 65536 + int(strRead[2])) / 1000
        self.lineEdit_36.setText("%.3f" % info2)

        # Z轴的终点坐标
        info3 = (int(strRead[5]) * 65536 + int(strRead[4])) / 1000
        self.lineEdit_37.setText("%.3f" % info3)

        # X轴的段点坐标
        info4 = (int(strRead[7]) * 65536 + int(strRead[6])) / 1000
        self.lineEdit_38.setText("%.3f" % info4)

        # Y轴的段点坐标
        info5 = (int(strRead[9]) * 65536 + int(strRead[8])) / 1000
        self.lineEdit_39.setText("%.3f" % info5)

        # Z轴的段点坐标
        info6 = (int(strRead[11]) * 65536 + int(strRead[10])) / 1000
        self.lineEdit_40.setText("%.3f" % info6)

        if int(strRead[12]) == 0:
            # 将 GUI 中的 lineEdit_44 文本框的内容设置为字符串 "0"。
            self.lineEdit_44.setText("0")

        if int(strRead[12]) == 1000:
            self.lineEdit_44.setText("1")

        read = master.execute(1, cst.HOLDING_REGISTERS, 0x2179, 2)
        print(read)
        strRead = [str(i) for i in read[0]]
        print(strRead)

        # 插补速度
        info7 = (int(strRead[1]) * 65536 + int(strRead[0])) / 1000
        self.lineEdit_34.setText("%.3f" % info7)

        read = master.execute(1, cst.HOLDING_REGISTERS, 0x216F, 6)
        print(read)
        strRead = [str(i) for i in read[0]]
        print(strRead)

        # X轴位置
        info8 = (int(strRead[1]) * 65536 + int(strRead[0])) / 1000
        self.s1__lb_60.setText("%.3f" % info8)

        # Y轴位置
        info9 = (int(strRead[3]) * 65536 + int(strRead[2])) / 1000
        self.s1__lb_61.setText("%.3f" % info9)

        # Z轴位置
        info10 = (int(strRead[5]) * 65536 + int(strRead[4])) / 1000
        self.s1__lb_62.setText("%.3f" % info10)

    # 用于切换到某种“手动操作”模式，在这种模式下，用户可能只需要看到verticalGroupBox_3组件，
    # 而其他组件（如参数设置、程序编辑和线 / 弧设置）则被隐藏。同时，窗口的标题也被更新为“手动操作”，以反映当前的操作模式。
    def Main_interface(self):    # 设置不同组件的可见性
        self.verticalGroupBox_3.setVisible(True)
        self.GroupBoxPara.setVisible(False)
        self.GroupBoxEditProgram.setVisible(False)
        self.GroupBoxLine_arc.setVisible(False)
        self.setWindowTitle(self._translate("Form", "手动操作"))

    def Basic_para_interface(self):
        self.verticalGroupBox_3.setVisible(False)
        self.GroupBoxPara.setVisible(True)
        self.GroupBoxEditProgram.setVisible(False)
        self.GroupBoxLine_arc.setVisible(False)
        self.setWindowTitle(self._translate("Form", "参数设置"))

    def peizhichuankou(self):
        # 调用self.ON_SetUp_Port()方法，并将返回值存储在vle变量中
        vle = self.ON_SetUp_Port()
        if vle == QDialog.Accepted:
            # readini将串口信息保存到配置文件
            # 在名为 "common" 的节下设置一个键为 "port"，值为 self.port 的键值对
            self.readini.set("common", "port", self.port)
            self.readini.set("common", "botelv", self.botelv)
            self.readini.set("common", "shujuwei", self.shujuwei)
            self.readini.set("common", "jiaoyanwei", self.jiaoyanwei)
            self.readini.set("common", "stop", self.stop)
            # 配置信息以键值对的形式保存，并最终写入到 self.conf_path 指定的文件中
            self.readini.write(open(self.conf_path, "w"))
        if vle == QDialog.Accepted:
            if self.ser is not None and self.ser.isOpen():
                self.stop_button.click()  #调用的是stop_button 按钮连接的信号槽  要记住
            self.open_button.click()
        # vle = self.ON_SetUp_Port()
        # global master
        # while True:
        #     if self.port != "":
        #         self.ser.port = self.port
        #         try:
        #             master = modbus_rtu.RtuMaster(serial.Serial(port=self.ser.port,
        #                                                         baudrate=int(self.botelv), bytesize=int(self.shujuwei),
        #                                                         parity=self.jiaoyanwei, stopbits=float(self.stop)))
        #             master.set_timeout(2,True)
        #             self.ser = master._serial
        #             self.s1__lb_42.setText(self._translate("Form", "当前状态：已连接控制器"))
        #             self.s1__lb_42.setFont(QFont("华文楷书", 10.5))
        #             self.s1__lb_42.setStyleSheet("color:rgb(250,0,0);")
        #
        #             self.s1__lb_43.setText(self._translate("Form", "已连接"))
        #             self.s1__lb_43.setFont(QFont("华文楷书", 10.5))
        #             self.s1__lb_43.setStyleSheet("color:rgb(250,0,0);")
        #
        #             self.readini.set("common", "port", self.port)
        #             self.readini.set("common", "botelv", self.botelv)
        #             self.readini.set("common", "shujuwei", self.shujuwei)
        #             self.readini.set("common", "jiaoyanwei", self.jiaoyanwei)
        #             self.readini.set("common", "stop", self.stop)
        #             self.readini.write(open(self.conf_path, "w"))
        #
        #         except Exception as err:
        #             QMessageBox.critical(self, "Port Error\n" + str(err), "此串口不能被打开！" + str(err))
        #
        #         if self.ser.isOpen():
        #             self.open_button.setEnabled(False)
        #             break
        #         else:
        #             vle = self.ON_SetUp_Port()
        #             print(vle)
        #             if vle != QDialog.Accepted:
        #                 break
        #
        #     else:
        #         vle = self.ON_SetUp_Port()
        #         print(vle)
        #         if vle != QDialog.Accepted:
        #             break

    def STOP(self):
        """
        停止按钮功能
        :return:
        """
        if self.timer1 > 0:
            # 停止计时器，并将timer1重置为0
            self.killTimer(self.timer1)
            self.timer1 = 0
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0010, output_value=0xff00))

        # 重启定时器，每100毫秒触发一次
        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def selectionchange(self):
        if self.s1__box_23.currentText() == "中文":
            self.Language = "CHINESE"
            self._trigger_zh_cn()

        if self.s1__box_23.currentText() == "ENGLISH":
            self.Language = "ENGLISH"
            self._trigger_english()

    def _trigger_english(self):
        print("[Form] Change to English")
        # 加载英文翻译文件
        self.trans.load("en_US.qm")
        _app = QApplication.instance()  # 获取app实例
        # 将翻译软件安装到应用程序中
        _app.installTranslator(self.trans)
        # 重新翻译界面
        self.retranslateUi(self)
        pass

    def _trigger_zh_cn(self):
        print("[Form] Change to zh_CN")
        # 通常 .ts文件是 Qt Linguist 编辑的翻译源文件，而 .qm文件是编译后的二进制翻译文件。
        self.trans.load("zh_CN.ts")
        _app = QApplication.instance()
        _app.installTranslator(self.trans)
        self.retranslateUi(self)
        pass

    def about(self):
        dialog = AboutFunc()
        # 显示并运行这个“关于”对话框。在 Qt 中，exec() 方法会阻塞当前线程，直到对话框关闭。
        dialog.exec()

    def New_program(self):
        # 创建对话框
        Dig = NewProgramDialogFunc()
        # 调用 Dig 对话框的 exec() 方法来显示它。如果对话框被用户接受（通常是点击了“确定”或“OK”按钮），则执行下面的代码块
        if Dig.exec() == QDialog.Accepted:
            # 获取对话框的数据
            self.ProjectName = Dig.getGongChengName()
            # 清空对象的 TableWidget中的所有内容
            self.TableWidget.clearContents()
            # 从 Dig 对话框中获取行数（可能是用户输入的），将其转换为整数，并设置到一个名为 lineEdit_28 的文本框中
            a = Dig.getHangNum()
            a = int(a)
            self.lineEdit_28.setText(str(a))
            # 初始化表格内容
            i=0
            while i< a:
                newItem = QTableWidgetItem(' ---请选择---  ')
                self.TableWidget.setItem(i,0,newItem)
                newItem2 = QTableWidgetItem(' ---请选择---  ')
                self.TableWidget.setItem(i, 1, newItem2)

                # 初始化表格，在 TableWidget 的第 i+1 行、第 3 列的位置插入一个包含数字 0 的新单元格项。
                newItem_3 = QTableWidgetItem(str(0))
                self.TableWidget.setItem(i, 2, newItem_3)

                newItem_4 = QTableWidgetItem(str(0))
                self.TableWidget.setItem(i, 3, newItem_4)

                newItem_5 = QTableWidgetItem(str(0))
                self.TableWidget.setItem(i, 4, newItem_5)

                newItem_6 = QTableWidgetItem(str(0))
                self.TableWidget.setItem(i, 5, newItem_6)

                newItem_7 = QTableWidgetItem(str(0))
                self.TableWidget.setItem(i, 6, newItem_7)

                i+=1

    def Delete_program(self):
        Dig = Delete_ProgramDialogFunc()
        if Dig.exec() == QDialog.Accepted:
            self.TableWidget.clearContents()
            self.lineEdit_28.setText('0')

    def eventFilter(self, a0: 'QObject', a1: 'QEvent'):
        # a0事件接收者，a1事件本身，从a0中获取的对象
        timerID = a0.property("type")
        if a1.type() == QEvent.MouseButtonDblClick:
            if timerID == 1:
                self.killTimer(self.timer1)
                self.timer1 = 0
            if timerID == 4:
                self.killTimer(self.timer2)
                self.timer2 = 0
        return super().eventFilter(a0,a1)

    def stopTimer1(self):
        if self.timer1>0:
            self.killTimer(self.timer1)
            self.timer1 = 0

    def startTimer1(self):
        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def stopTimer2(self):
        if self.timer2 > 0:
            self.killTimer(self.timer2)
            self.timer2 = 0

    def startTimer2(self):
        if self.timer2==0:
            self.timer2 = self.startTimer(100)

    def Manual_input_show(self):  # 定义的手动操作子区域
        self.verticalGroupBox_3.setVisible(True)
        self.GroupBoxPara.setVisible(False)
        self.GroupBoxEditProgram.setVisible(False)
        self.GroupBoxLine_arc.setVisible(False)
        self.setWindowTitle(self._translate("Form", "手动操作"))

        self.refresh_manual_data()
        self.Only_two()
        if self.timer2 > 0:
            self.killTimer(self.timer2)
            self.timer2 = 0
        if self.timer1 == 0:
            self.timer1 = self.startTimer(100)

    def refresh_manual_data(self):

        # todo 因为要将从Modbus Slave 读取到的数据放到手动操作的各个文本框中，而现在读取到的数据在表格中，并且分割了五份，
        # todo 要将读取到的数据分别放到5个文本框和勾选输入口状态和输出口状态
        # todo 在Modbus Slave中写寄存器数据的时候
        # todo 前0个寄存器是对应的“上电自动运行”
        # todo 第1个寄存器到第10个寄存器数据对应的是 5个文本框
        # todo 第11个寄存器到第14个寄存器数据对应的是 “输出口选择”
        # todo 第15个寄存器到第21个寄存器数据对应的是 “输入口状态”
        # todo 第22个寄存器到第25个寄存器数据对应的是 “输出口状态”
        # todo 第26个寄存器到第28个寄存器数据对应的是 “电机状态”
        # 读取 Modbus 数据
        try:
            read = master.execute(1, cst.HOLDING_REGISTERS, 0x216F, 45)
            print(read)
        except Exception as e:
            return

        strRead = [i for i in read]

        if strRead is None:
            QMessageBox.critical(self, "Port Error\n", "暂时没有工程参数！")
            return

        #X轴位置
        # if int(strRead[0]) > 0x7FFF:
        #     strRead[0] = int(strRead[0]) - 0x10000
        if not isinstance(strRead, (list, tuple)): # 如果不是list和tuple,就返回
            return

        # 从strRead列表中取出索引为1的元素，并将其值赋给变量_readOne
        """Modbus协议中常常使用两个连续的16位寄存器来表示一个32位有符号整数。
        当高位寄存器（即这里的strRead[1]）的值大于32767时，表示这是一个负数，并且需要减去65536以得到正确的值。"""
        _readOne = strRead[1]
        if _readOne > 32767:
            strRead[1] = _readOne - 65536

        # if int(strRead[2]) > 0x7FFF:
        #     strRead[2] = int(strRead[2]) - 0x10000

        _readOne = strRead[3]
        if _readOne > 32767:
            strRead[3] = _readOne - 65536

        # if int(strRead[4]) > 0x7FFF:
        #     strRead[4] = int(strRead[4]) - 0x10000

        _readOne = strRead[5]
        if _readOne > 32767:
            strRead[5] = _readOne - 65536

        # X轴位置
        info3 = (strRead[1] * 65536 + strRead[0]) / 1000
        print(info3)
        self.s1__lb_44.setText("%.3f" % info3)  # X轴位置文本框

        # Y轴位置
        info4 = (strRead[3] * 65536 + strRead[2]) / 1000
        self.s1__lb_45.setText("%.3f" % info4)  # Y轴位置文本框

        # Z轴位置
        info5 = (strRead[5] * 65536 + strRead[4]) / 1000
        self.s1__lb_46.setText("%.3f" % info5)  # Z轴位置文本框

        #手动速度
        info1 = (strRead[11] * 65536 + strRead[10]) / 1000
        self.lineEdit_4.setText("%.3f" % info1)  # 手动速度文本框

        #步进距离
        info2 = (strRead[13] * 65536 + strRead[12]) / 1000
        self.lineEdit_5.setText("%.3f" % info2)  # 步进文本框

        if (strRead[14]== 1):  # 对应上电自动运行勾上
            self.CheckBox_19.setChecked(True)
        if (strRead[14]== 0):  # 对应上电自动运行勾掉
            self.CheckBox_19.setChecked(False)

        # 对应 电机状态 三个寄存器
        if (strRead[15]== 1):  # 对应X轴电机状态寄存器勾上
            self.CheckBox_16.setChecked(True)
        if (strRead[15]== 0):  # 对应X轴电机状态寄存器勾掉
            self.CheckBox_16.setChecked(False)

        if (strRead[16] == 1):  # 对应Y轴电机状态寄存器勾上
            self.CheckBox_17.setChecked(True)
        if (strRead[16] == 0):  # 对应Y轴电机状态寄存器勾掉
            self.CheckBox_17.setChecked(False)

        if (strRead[17] == 1):  # 对应Z轴电机状态寄存器勾上
            self.CheckBox_18.setChecked(True)
        if (strRead[17] == 0):  # 对应Z轴电机状态寄存器勾掉
            self.CheckBox_18.setChecked(False)

        # 对应输入口状态 七个寄存器
        if (strRead[20] == 1):  # 对应输入口状态IN1寄存器勾上
            self.CheckBox_5.setChecked(True)
        if (strRead[20] == 0):  # 对应输入口状态IN1寄存器勾掉
            self.CheckBox_5.setChecked(False)

        if (strRead[21] == 1):  # 对应输入口状态IN2寄存器勾上
            self.CheckBox_6.setChecked(True)
        if (strRead[21] == 0):  # 对应输入口状态IN2寄存器勾掉
            self.CheckBox_6.setChecked(False)

        if (strRead[22] == 1):  # 对应输入口状态IN3寄存器勾上
            self.CheckBox_7.setChecked(True)
        if (strRead[22] == 0):  # 对应输入口状态IN3寄存器勾掉
            self.CheckBox_7.setChecked(False)

        if (strRead[23] == 1):  # 对应输入口状态IN4寄存器勾上
            self.CheckBox_8.setChecked(True)
        if (strRead[23] == 0):  # 对应输入口状态IN4寄存器勾掉
            self.CheckBox_8.setChecked(False)

        if (strRead[28] == 1):  # 对应输入口状态ORG1寄存器勾上
            self.CheckBox_9.setChecked(True)
        if (strRead[28] == 0):  # 对应输入口状态ORG1寄存器勾掉
            self.CheckBox_9.setChecked(False)

        if (strRead[29] == 1):  # 对应输入口状态ORG2寄存器勾上
            self.CheckBox_10.setChecked(True)
        if (strRead[29] == 0):  # 对应输入口状态ORG2寄存器勾掉
            self.CheckBox_10.setChecked(False)

        if (strRead[30] == 1):  # 对应输入口状态ORG3寄存器勾上
            self.CheckBox_11.setChecked(True)
        if (strRead[30] == 0):  # 对应输入口状态ORG3寄存器勾掉
            self.CheckBox_11.setChecked(False)

        # 对应“输出口状态”四个寄存器
        if (strRead[33] == 1):  # 对应输出口状态OUT1寄存器勾上
            self.CheckBox_12.setChecked(True)
        if (strRead[33] == 0):  # 对应输出口状态OUT1寄存器勾掉
            self.CheckBox_12.setChecked(False)

        if (strRead[34] == 1):  # 对应输出口状态OUT2寄存器勾上
            self.CheckBox_13.setChecked(True)
        if (strRead[34] == 0):  # 对应输出口状态OUT2寄存器勾掉
            self.CheckBox_13.setChecked(False)

        if (strRead[35] == 1):  # 对应输出口状态OUT3寄存器勾上
            self.CheckBox_14.setChecked(True)
        if (strRead[35] == 0):  # 对应输出口状态OUT3寄存器勾掉
            self.CheckBox_14.setChecked(False)

        if (strRead[36] == 1):  # 对应输入口状态OUT4寄存器勾上
            self.CheckBox_15.setChecked(True)
        if (strRead[36] == 0):  # 对应输入口状态OUT4寄存器勾掉
            self.CheckBox_15.setChecked(False)

        # 手动操作——1表示为步进开
        read = master.execute(1, cst.HOLDING_REGISTERS, 0x21A1, 1)
        print(read)
        strRead = [str(i) for i in read]
        # strRead = [self.convertHex(i) for i in read]
        print(strRead)

        if (int(strRead[0]) == 1):
            self.CheckBox_20.setChecked(True)
        if (int(strRead[0]) == 0):
            self.CheckBox_20.setChecked(False)

    def Only_two(self):
        # 上电是否自动运行状态
        read = master.execute(1, cst.HOLDING_REGISTERS, 0x217D, 1)
        print(read)
        strRead = [str(i) for i in read]
        # strRead = [self.convertHex(i) for i in read]
        print(strRead)
        if (int(strRead[0]) == 1):  # 对应上电自动运行
            self.CheckBox_19.setChecked(True)

        if (int(strRead[0]) == 0):  # 对应上电自动运行
            self.CheckBox_19.setChecked(False)

        # 手动操作
        read = master.execute(1, cst.HOLDING_REGISTERS, 0x21A1, 1)
        print(read)
        strRead = [str(i) for i in read]
        # strRead = [self.convertHex(i) for i in read]
        print(strRead)

        if (int(strRead[0]) == 1):
            self.CheckBox_20.setChecked(True)
        if (int(strRead[0]) == 0):  # 对应上电自动运行
            self.CheckBox_20.setChecked(False)
        #————————————刷新到此结束————————————————————————

    def set_qcheck_state1(self,state):
        """
        输出口选择OUT1
        :param state:
        :return:
        """
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        self.CheckBox_1 = self.sender()

        if state == QtCore.Qt.Checked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0011, output_value=0xff00))

        elif state == QtCore.Qt.Unchecked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0011, output_value=0x0000))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def set_qcheck_state2(self,state):
        """
        输出口选择OUT2
        :param state:
        :return:
        """
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        self.CheckBox_2 = self.sender()

        if state == QtCore.Qt.Checked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0012, output_value=0xff00))

        elif state == QtCore.Qt.Unchecked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0012, output_value=0x0000))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def set_qcheck_state3(self,state):
        """
        输出口选择OUT3
        :return:
        """
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        self.CheckBox_3 = self.sender()

        if state == QtCore.Qt.Checked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0013, output_value=0xff00))

        elif state == QtCore.Qt.Unchecked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0013, output_value=0x0000))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def set_qcheck_state4(self,state):
        """
        输出口选择OUT4
        :param state:
        :return:
        """
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        self.CheckBox_4 = self.sender()

        if state == QtCore.Qt.Checked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0014, output_value=0xff00))

        elif state == QtCore.Qt.Unchecked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0014, output_value=0x0000))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def set_qcheck_state5(self,state):
        """
        上电自动运行
        :return:
        """
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器

        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0

        self.CheckBox_19 = self.sender()
        if state == QtCore.Qt.Unchecked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_REGISTER, 0x217D, output_value=0))

        elif state == QtCore.Qt.Checked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("unconnected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_REGISTER, 0x217D, output_value=1))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def set_qcheck_state6(self,state):
        """
        步进前边的勾选框
        :return:
        """
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器

        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        self.CheckBox_20= self.sender()
        if state == QtCore.Qt.Checked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x001D, output_value=0xff00))

        elif state == QtCore.Qt.Unchecked:
            logger = modbus_tk.utils.create_logger("console")
            logger.info("connected")
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x001D, output_value=0x0000))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def Para_save(self):
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器

        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0022, output_value=0xff00))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def minus_down(self):
        """
        X-按钮功能
        :return:
        """
        if self.timer1>0:
            self.killTimer(self.timer1)
            self.timer1 = 0

        global master
        # self.lineEdit_10 = self.sender()  qt  sender
        button = self.sender()  # 判断是由哪个控件发出的信号  可以更改 button=obj
        name = button.property("name")
        method = button.property("method")
        circle = 0x0001
        if name == "X" and method =="plus":
            circle = 0x0002

        if name == "Y" and method == "sub":
            circle = 0x0004
        if name == "Y" and method == "plus":
            circle = 0x0005

        if name == "Z" and method == "sub":
            circle = 0x0007
        if name == "Z" and method == "plus":
            circle = 0x0008

        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, circle, output_value=0xff00))

    def plus_up(self):
        """
        X-按钮功能
        :return:
        """

        global master
        button = self.sender()
        name = button.property("name")
        method = button.property("method")
        circle = 0x0001
        if name == "X" and method == "plus":
            circle = 0x0002

        if name == "Y" and method == "sub":
            circle = 0x0004
        if name == "Y" and method == "plus":
            circle = 0x0005

        if name == "Z" and method == "sub":
            circle = 0x0007
        if name == "Z" and method == "plus":
            circle = 0x0008


        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, circle, output_value=0x0000))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def Back_X_zero(self):
        """
        X轴回零按钮功能
        :return:
        """
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0003, output_value=0xff00))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def Back_Y_zero(self):
        """
        Y轴回零按钮功能
        :return:
        """
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0006, output_value=0xff00))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def Back_Z_zero(self):
        """
        Z回零按钮功能
        :return:
        """
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0009, output_value=0xff00))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def XY_Line_interpolation(self):
        """
        XY直线插补命令
        :return:
        """
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x001E, output_value=0xff00))

    def XZ_Line_interpolation(self):
        """
        XZ直线插补命令
        :return:
        """
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x001F, output_value=0xff00))

    def YZ_Line_interpolation(self):
        """
        YZ直线插补命令
        :return:
        """
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0027, output_value=0xff00))

    def XY_Arc_interpolation(self):
        """
        XY圆弧插补命令
        :return:
        """
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0024, output_value=0xff00))

    def XZ_Arc_interpolation(self):
        """
        XZ圆弧插补命令
        :return:
        """
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0025, output_value=0xff00))

    def YZ_Arc_interpolation(self):
        """
        YZ圆弧插补命令
        :return:
        """
        global master
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_SINGLE_COIL, 0x0026, output_value=0xff00))

    def Enter_speed(self):  # 定义的“手动速度”文本框后面的“确定”按钮功能
        self.verticalGroupBox_3.setVisible(True)
        self.GroupBoxPara.setVisible(False)
        self.GroupBoxEditProgram.setVisible(False)
        # todo 在这个位置修改程序，读取数据时只读取手动操作界面的数据
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        # 向Modbus Slave寄存器写数据
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")

        a_1 = float(self.lineEdit_4.text())
        b_1 = a_1 * 1000
        c_1 = int(b_1 % 65536)
        print(c_1)
        d_1 = int(b_1 // 65536)
        print(d_1)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2179, output_value=[c_1, d_1]))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def Enter_bujin(self):  # 定义的“步进”文本框后面的“确定”按钮功能
        self.verticalGroupBox_3.setVisible(True)
        self.GroupBoxPara.setVisible(False)
        self.GroupBoxEditProgram.setVisible(False)
        # todo 在这个位置修改程序，读取数据时只读取手动操作界面的数据
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        # 向Modbus Slave寄存器写数据
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        # todo  4.3号任务：解决目前这个只能下发整数的情况
        a_2 = float(self.lineEdit_5.text())
        b_2 = a_2 * 1000
        c_2 = int(b_2 % 65536)
        print(c_2)
        d_2 = int(b_2 // 65536)
        print(d_2)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x217B, output_value=[c_2, d_2]))

        if self.timer1==0:
            self.timer1 = self.startTimer(100)

    def Parameter_setting_show(self):  # 定义的“参数设置”按钮功能
        self.verticalGroupBox_3.setVisible(False)
        self.GroupBoxPara.setVisible(True)
        self.GroupBoxEditProgram.setVisible(False)
        self.GroupBoxLine_arc.setVisible(False)
        self.setWindowTitle(self._translate("Form", "参数设置"))
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0

    def Parameter_Read(self):
        """
        “参数读取”按钮功能
        :return:
        """

        self.Parameter_Read_data()
        # if self.timer1 > 0:
        #     self.killTimer(self.timer1)
        #     self.timer1 = 0
        # if self.timer3 > 0:
        #     self.killTimer(self.timer3)
        #     self.timer3 = 0
        # if self.timer2 == 0:
        #     self.timer2 = self.startTimer(100)

    def Parameter_Read_data(self):

        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        # self.s1__box_9.currentText().strip()
        # 读保持寄存器
        read = master.execute(1, cst.HOLDING_REGISTERS, 0x2000, 39)  # 这里可以修改需要读取的功能码
        print(read)
        strRead = [str(i) for i in read]
        # strRead = [self.convertHex(i) for i in read]
        print(strRead)
        # todo 因为要将从Modbus Slave 读取到的数据放到手动操作的各个文本框中，而现在读取到的数据在表格中，并且分割了五份，
        # 所以就需要在目前程序的基础上修改程序，
        # todo 只需要添加一下放置数据的位置和将数据分割的再细一点
        # todo 将寄存器里边从第8个数据开始放置到参数设置界面
        # info6---info21分别对应参数设置各个文本框
        # 寄存器41----46对应系统加速度文本框数据
        #导程X
        info9 = (int(strRead[3]) * 65536 + int(strRead[2])) / 1000
        self.lineEdit_13.setText("%.3f" % info9)
        #细分数X
        info12 = int(strRead[5]) * 65536 + int(strRead[4])
        self.lineEdit_16.setText(str(info12))
        #行程X
        info15 = (int(strRead[7]) * 65536 + int(strRead[6])) / 1000
        self.lineEdit_19.setText("%.3f" % info15)
        #回原点方向X
        L_1 = int(strRead[8])
        if L_1 == 1:
            self.s1__box_20.setCurrentText("Y")
            self.s1__box_20.setFont(QFont("Times New Roman", 8))
        if L_1 == 0:
            self.s1__box_20.setCurrentText("N")
            self.s1__box_20.setFont(QFont("Times New Roman", 8))
        #回原点速度文本框X
        info18 = (int(strRead[10]) * 65536 + int(strRead[9])) / 1000
        self.lineEdit_22.setText("%.3f" % info18)
        #原点回退距离X
        info21 = (int(strRead[12]) * 65536 + int(strRead[11])) / 1000
        self.lineEdit_25.setText("%.3f" % info21)
        #脉冲当量X
        if info9 != 0 and info12 != 0:
            v_x = info9 / info12
            v_x = ("%.3f" % v_x)
            print(v_x)
            self.lineEdit_10.setText(str(v_x))

        #导程Y
        info10 = (int(strRead[16]) * 65536 + int(strRead[15])) / 1000
        self.lineEdit_14.setText("%.3f" % info10)
        #细分数Y
        info13 = int(strRead[18]) * 65536 + int(strRead[17])
        self.lineEdit_17.setText(str(info13))
        #行程Y
        info16 = (int(strRead[20]) * 65536 + int(strRead[19])) / 1000
        self.lineEdit_20.setText("%.3f" % info16)
        #回原点方向Y
        Y_1 = int(strRead[21])
        if Y_1 == 1:
            self.s1__box_21.setCurrentText("Y")
            self.s1__box_21.setFont(QFont("Times New Roman", 8))
        if Y_1 == 0:
            self.s1__box_21.setCurrentText("N")
            self.s1__box_21.setFont(QFont("Times New Roman", 8))
        #回原点速度Y
        info19 = (int(strRead[23]) * 65536 + int(strRead[22])) / 1000
        self.lineEdit_23.setText("%.3f" % info19)
        #原点回退距离Y
        info22 = (int(strRead[25]) * 65536 + int(strRead[24])) / 1000
        self.lineEdit_26.setText("%.3f" % info22)
        #脉冲当量Y
        if info10 != 0 and info13 != 0:
            v_y = info10 / info13
            v_y = ("%.3f" % v_y)
            print(v_y)
            self.lineEdit_11.setText(str(v_y))

        #导程Z
        info11 = (int(strRead[29]) * 65536 + int(strRead[28])) / 1000
        self.lineEdit_15.setText("%.3f" % info11)
        #细分数Z
        info14 = int(strRead[31]) * 65536 + int(strRead[30])
        self.lineEdit_18.setText(str(info14))
        #行程Z
        info17 = (int(strRead[33]) * 65536 + int(strRead[32])) / 1000
        self.lineEdit_21.setText("%.3f" % info17)
        #回原点方向Z
        X_1 = int(strRead[34])
        if X_1 == 1:
            self.s1__box_22.setCurrentText("Y")
            self.s1__box_22.setFont(QFont("Times New Roman", 8))
        if X_1 == 0:
            self.s1__box_22.setCurrentText("N")
            self.s1__box_22.setFont(QFont("Times New Roman", 8))
        #回原点速度Z
        info20 = (int(strRead[36]) * 65536 + int(strRead[35])) / 1000
        self.lineEdit_24.setText("%.3f" % info20)
        #原点回退距离Z
        info23 = (int(strRead[38]) * 65536 + int(strRead[37])) / 1000
        self.lineEdit_27.setText("%.3f" % info23)
        #脉冲当量Z
        if info11 != 0 and info14 != 0:
            v_z = info11 / info14
            v_z = ("%.3f" % v_z)
            self.lineEdit_12.setText(str(v_z))

        read = master.execute(1, cst.HOLDING_REGISTERS, 0x2165, 10)  # 这里可以修改需要读取的功能码
        print(read)
        strRead = [str(i) for i in read]
        # strRead = [self.convertHex(i) for i in read]
        print(strRead)
        #系统加速度X
        info6 = (int(strRead[1]) * 65536 + int(strRead[0])) / 1000
        self.lineEdit_9.setText("%.3f" % info6)
        #系统加速度Y
        info7 = (int(strRead[3]) * 65536 + int(strRead[2])) / 1000
        self.lineEdit_29.setText("%.3f" % info7)
        #系统加速度Z
        info8 = (int(strRead[5]) * 65536 + int(strRead[4])) / 1000
        self.lineEdit_30.setText("%.3f" % info8)

    # todo 在这个位置添加参数写入函数
    def Parameter_Write(self):
        """
        “参数写入”按钮功能
        :return:
        """
        # todo 在这个位置修改程序，读取数据时只读取手动操作界面的数据
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        # 向Modbus Slave寄存器写数据
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        # todo  4.3号任务：解决目前这个只能下发整数的情况

        a_3 = float(self.lineEdit_9.text())  # X轴系统加速度
        b_3 = a_3 * 1000
        c_3 = int(b_3 % 65536)
        print(c_3)
        d_3 = int(b_3 // 65536)
        print(d_3)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2165, output_value=[c_3, d_3]))

        a_4 = float(self.lineEdit_29.text())  # Y轴系统加速度
        b_4 = a_4 * 1000
        c_4 = int(b_4 % 65536)
        print(c_4)
        d_4 = int(b_4 // 65536)
        print(d_4)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2167, output_value=[c_4, d_4]))

        a_5 = float(self.lineEdit_30.text())  # Z轴系统加速度
        b_5 = a_5 * 1000
        c_5 = int(b_5 % 65536)
        print(c_5)
        d_5 = int(b_5 // 65536)
        print(d_5)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2169, output_value=[c_5, d_5]))

        a_6 = float(self.lineEdit_13.text())  # 导程X
        b_6 = a_6 * 1000
        c_6 = int(b_6 % 65536)
        print(c_6)
        d_6 = int(b_6 // 65536)
        print(d_6)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2002, output_value=[c_6, d_6]))

        a_9 = int(self.lineEdit_16.text())  # 细分数X
        b_9 = a_9
        c_9 = int(b_9 % 65536)
        print(c_9)
        d_9 = int(b_9 // 65536)
        print(d_9)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2004, output_value=[c_9, d_9]))

        a_12 = float(self.lineEdit_19.text())  # 行程X
        b_12 = a_12 * 1000
        c_12 = int(b_12 % 65536)
        print(c_12)
        d_12 = int(b_12 // 65536)
        print(d_12)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2006, output_value=[c_12, d_12]))

        if self.s1__box_20.currentText() == "Y":
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2008, output_value=[1]))

        if self.s1__box_20.currentText() == "N":
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2008, output_value=[0]))

        a_15 = float(self.lineEdit_22.text())  # 回原点速度X
        b_15 = a_15 * 1000
        c_15 = int(b_15 % 65536)
        print(c_15)
        d_15 = int(b_15 // 65536)
        print(d_15)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2009, output_value=[c_15, d_15]))

        a_18 = float(self.lineEdit_25.text())  # 原点回退距X
        b_18 = a_18 * 1000
        c_18 = int(b_18 % 65536)
        print(c_18)
        d_18 = int(b_18 // 65536)
        print(d_18)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x200B, output_value=[c_18, d_18]))

        a_7 = float(self.lineEdit_14.text())  # 导程Y
        b_7 = a_7 * 1000
        c_7 = int(b_7 % 65536)
        print(c_7)
        d_7 = int(b_7 // 65536)
        print(d_7)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,0x200F, output_value=[c_7, d_7]))

        a_10 = int(self.lineEdit_17.text())  # 细分数Y
        b_10 = a_10
        c_10 = int(b_10 % 65536)
        print(c_10)
        d_10 = int(b_10 // 65536)
        print(d_10)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2011, output_value=[c_10, d_10]))

        a_13 = float(self.lineEdit_20.text())  # 行程Y
        b_13 = a_13 * 1000
        c_13 = int(b_13 % 65536)
        print(c_13)
        d_13 = int(b_13 // 65536)
        print(d_13)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2013, output_value=[c_13, d_13]))

        if self.s1__box_21.currentText() == "Y":
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2015, output_value=[1]))

        if self.s1__box_21.currentText() == "N":
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2015, output_value=[0]))

        a_16 = float(self.lineEdit_23.text())  # 回原点速度Y
        b_16 = a_16 * 1000
        c_16 = int(b_16 % 65536)
        print(c_16)
        d_16 = int(b_16 // 65536)
        print(d_16)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2016, output_value=[c_16, d_16]))

        a_19 = float(self.lineEdit_26.text())  # 原点回退距Y
        b_19 = a_19 * 1000
        c_19 = int(b_19 % 65536)
        print(c_19)
        d_19 = int(b_19 // 65536)
        print(d_19)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2018, output_value=[c_19, d_19]))

        a_8 = float(self.lineEdit_15.text())  # 导程Z
        b_8 = a_8 * 1000
        c_8 = int(b_8 % 65536)
        print(c_8)
        d_8 = int(b_8 // 65536)
        print(d_8)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x201C, output_value=[c_8, d_8]))

        a_11 = int(self.lineEdit_18.text())  # 细分数Z
        b_11 = a_11
        c_11 = int(b_11 % 65536)
        print(c_11)
        d_11 = int(b_11 // 65536)
        print(d_11)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x201E, output_value=[c_11, d_11]))

        a_14 = float(self.lineEdit_21.text())  # 行程Z
        b_14 = a_14 * 1000
        c_14 = int(b_14 % 65536)
        print(c_14)
        d_14 = int(b_14 // 65536)
        print(d_14)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2020, output_value=[c_14, d_14]))

        if self.s1__box_22.currentText() == "Y":
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2022, output_value=[1]))

        if self.s1__box_22.currentText() == "N":
            for i in range(1):
                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2022, output_value=[0]))

        a_17 = float(self.lineEdit_24.text())  # 回原点速度Z
        b_17 = a_17 * 1000
        c_17 = int(b_17 % 65536)
        print(c_17)
        d_17 = int(b_17 // 65536)
        print(d_17)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2023, output_value=[c_17, d_17]))

        a_20 = float(self.lineEdit_27.text())  # 原点回退距Z
        b_20 = a_20 * 1000
        c_20 = int(b_20 % 65536)
        print(c_20)
        d_20 = int(b_20 // 65536)
        print(d_20)
        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x2025, output_value=[c_20, d_20]))
        return

    def Editing_program_show(self): # 定义的程序编辑子区域
        self.verticalGroupBox_3.setVisible(False)
        self.GroupBoxPara.setVisible(False)
        self.GroupBoxEditProgram.setVisible(True)
        self.GroupBoxLine_arc.setVisible(False)
        self.setWindowTitle(self._translate("Form", "程序编辑"))
        self.TableWidget.clearContents()
        self.Editing_program_show_data()  #调用所要刷新数据的函数即可
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0
        # if self.timer2 > 0:
        #     self.killTimer(self.timer2)
        #     self.timer2 = 0
        # if self.timer3 == 0:
        #     self.timer3 = self.startTimer(1000)

    def Editing_program_show_data(self):
        self.TableWidget.clearContents()
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        global master
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        read = master.execute(1, cst.HOLDING_REGISTERS, 0x219C, 1)   # 这里可以修改需要读取的功能码  0x0090改为0x219C
        print(type(read))
        print('read= ',read)
        strRead = [str(i) for i in read]
        # strRead = [str(i) for i in read]
        print(strRead)

        A_1 = int(strRead[0])  # A_1 表示的是要读取的工程参数的行数
        print(A_1)
        self.lineEdit_28.setText(strRead[0])

        B_1 = A_1 * 12  # 每行12个寄存器 B_1 = A_1 * 12 表示要读取的寄存器总数
        print(B_1)

        C_1 = A_1 // 10  # C_1表示的整10行数
        print(C_1)

        D_1 = A_1 % 10  # 行数的个位数
        print(D_1)

        E_1 = C_1 + 1  # E_1是读取的次数
        print(E_1)

        if A_1 == 0:
            QMessageBox.critical(self, "Port Error\n", "暂时没有工程参数！")
        if A_1 != 0:
            if E_1 == 1:  # 已修改（成功）
                i = 1
                while i <= D_1:
                    read = master.execute(1, cst.HOLDING_REGISTERS, 0x0000 + (i - 1) * 12 ,
                                          12)  # 这里可以加一个判断  如果是1  就会显示“流程控制”
                    print(read)
                    strRead = [str(i) for i in read]
                    # strRead = [self.convertHex(i) for i in read]
                    print(strRead)

                    F_1 = int(strRead[0])
                    G_1 = int(strRead[1])
                    H_1 = float((self.handleInt2(strRead[3],strRead[2])) / 1000)
                    print(H_1)
                    I_1 = float((self.handleInt2(strRead[5],strRead[4])) / 1000)
                    print(I_1)
                    J_1 = float((self.handleInt2(strRead[7],strRead[6])) / 1000)
                    print(J_1)
                    K_1 = float((self.handleInt2(strRead[9],strRead[8])) / 1000)
                    print(K_1)
                    L_1 = float((self.handleInt2(strRead[11],strRead[10])) / 1000)
                    print(L_1)
#todo 由于数据类型的要求，看看下边的数据可不可以先不放置，仅仅是因为数据类型有错误
#——————————————————————————————————————
                    comboBoxList1 = [self._translate("Form", "系统操作"), self._translate("Form", "流程控制"),
                                     self._translate("Form", "输出口设置"), self._translate("Form", "回零运动"),
                                     self._translate("Form", "插补运动"), self._translate("Form", "独立运动")]
                    comboBoxList2 = {
                        self._translate("Form", "系统操作"): [self._translate("Form", "停止"), self._translate("Form", "启动"),
                                                          self._translate("Form", "暂停"), self._translate("Form", "恢复"),
                                                          self._translate("Form", "延时等待"),
                                                          self._translate("Form", "等待电机完成"),
                                                          self._translate("Form", "停止电机运动"),
                                                          self._translate("Form", "常等待")],
                        self._translate("Form", "流程控制"): [self._translate("Form", "程序间跳转"),
                                                          self._translate("Form", "程序循环"),
                                                          self._translate("Form", "输入跳转"),
                                                          self._translate("Form", "开启输入中断"),
                                                          self._translate("Form", "关闭输入中断")],
                        self._translate("Form", "输出口设置"): [self._translate("Form", "输出口设置")],
                        self._translate("Form", "回零运动"): [self._translate("Form", "设置回零速度"),
                                                          self._translate("Form", "启动回零")],
                        self._translate("Form", "插补运动"): [self._translate("Form", "设置点位速度"),
                                                          self._translate("Form", "三轴相对运动"),
                                                          self._translate("Form", "单轴绝对运动"),
                                                          self._translate("Form", "XY绝对运动"),
                                                          self._translate("Form", "XZ绝对运动"),
                                                          self._translate("Form", "YZ绝对运动"),
                                                          self._translate("Form", "三轴绝对运动"),
                                                          self._translate("Form", "XY圆弧插补"),
                                                          self._translate("Form", "XZ圆弧插补"),
                                                          self._translate("Form", "YZ圆弧插补")],
                        self._translate("Form", "独立运动"): [self._translate("Form", "独立运动速度"),
                                                          self._translate("Form", "相对运动"),
                                                          self._translate("Form", "X绝对运动"),
                                                          self._translate("Form", "Y绝对运动"),
                                                          self._translate("Form", "Z绝对运动")]}

                    if F_1 == 1:  #已修改完成
                        if G_1 == 1:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][0]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)

                            # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                            # 在指令的下拉框中设置“停止”
                        elif G_1 == 2:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][1]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)

                            # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                            # 在指令的下拉框中设置“启动”
                        elif G_1 == 3:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][2]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)

                            # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                            # 在指令的下拉框中设置“暂停”
                        elif G_1 == 4:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][3]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)

                            # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                            # 在指令的下拉框中设置“恢复”
                        elif G_1 == 5:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][4]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            if H_1>=0:
                                newItem_1 = QTableWidgetItem(str(H_1))
                                self.TableWidget.setItem(i - 1, 2, newItem_1)
                            else:
                                QMessageBox.critical(self, "Port Error\n", "1_5 参数1数据错误！")
                                return None

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)

                            # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                            # 在指令的下拉框中设置“延时等待”
                        elif G_1 == 6:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][5]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            I_1 = int(I_1)
                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            J_1 = int(J_1)
                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 7:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][6]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            I_1 = int(I_1)
                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            J_1 = int(J_1)
                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 8:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "系统操作")][7]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)
                            newItem_1 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                    elif F_1 == 2:
                        if G_1 == 1:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "流程控制")][0]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 2:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "流程控制")][1]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            I_1 = int(I_1)
                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 3:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "流程控制")][2]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            I_1 = int(I_1)
                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            J_1 = int(J_1)
                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 4:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "流程控制")][3]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            I_1 = int(I_1)
                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            J_1 = int(J_1)
                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 5:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "流程控制")][4]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                    elif F_1 == 3:
                        if G_1 == 1:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "输出口设置")][0]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            I_1 = int(I_1)
                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                    elif F_1 == 4:
                        if G_1 == 1:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "回零运动")][0]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 2:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "回零运动")][1]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            I_1 = int(I_1)
                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            J_1 = int(J_1)
                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                    elif F_1 == 5:
                        if G_1 == 1:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][0]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 2:
                            print()
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][1]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 3:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][2]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 4:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][3]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 5:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][4]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 6:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][5]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 7:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][6]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 8:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][7]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(K_1))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            L_1 = int(L_1)
                            newItem_5 = QTableWidgetItem(str(L_1))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 9:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][8]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(K_1))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            L_1 = int(L_1)
                            newItem_5 = QTableWidgetItem(str(L_1))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 10:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "插补运动")][9]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(K_1))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            L_1 = int(L_1)
                            newItem_5 = QTableWidgetItem(str(L_1))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                    elif F_1 == 6:
                        if G_1 == 1:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "独立运动")][0]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            H_1 = int(H_1)
                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 2:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "独立运动")][1]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(I_1))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(J_1))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 3:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "独立运动")][2]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 4:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "独立运动")][3]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)
                        elif G_1 == 5:
                            op = comboBoxList1[F_1 - 1]
                            newItem_6 = QTableWidgetItem(str(op))
                            self.TableWidget.setItem(i - 1, 0, newItem_6)

                            ob = comboBoxList2[self._translate("Form", "独立运动")][4]
                            newItem_7 = QTableWidgetItem(str(ob))
                            self.TableWidget.setItem(i - 1, 1, newItem_7)

                            newItem_1 = QTableWidgetItem(str(H_1))
                            self.TableWidget.setItem(i - 1, 2, newItem_1)

                            newItem_2 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 3, newItem_2)

                            newItem_3 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 4, newItem_3)

                            newItem_4 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 5, newItem_4)

                            newItem_5 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(i - 1, 6, newItem_5)

                            i += 1
                    if F_1 < 1 or F_1 > 6:
                        newItem_100 = QTableWidgetItem("无")
                        self.TableWidget.setItem(i - 1, 0, newItem_100)

                        newItem_101 = QTableWidgetItem("无")
                        self.TableWidget.setItem(i - 1, 1, newItem_101)

                        newItem_102 = QTableWidgetItem(str(0))
                        self.TableWidget.setItem(i - 1, 2, newItem_102)

                        newItem_103 = QTableWidgetItem(str(0))
                        self.TableWidget.setItem(i - 1, 3, newItem_103)

                        newItem_104 = QTableWidgetItem(str(0))
                        self.TableWidget.setItem(i - 1, 4, newItem_104)

                        newItem_105 = QTableWidgetItem(str(0))
                        self.TableWidget.setItem(i - 1, 5, newItem_105)

                        newItem_106 = QTableWidgetItem(str(0))
                        self.TableWidget.setItem(i - 1, 6, newItem_106)
                    i+=1

            if E_1 >= 2:
                if D_1 == 0:
                    i = 1
                    while i <= C_1:
                        j = 1
                        while j <= 10:
                            read = master.execute(1, cst.HOLDING_REGISTERS, 0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                  12)
                            print(read)
                            strRead = [str(i) for i in read]
                            # strRead = [self.convertHex(i) for i in read]
                            print(strRead)

                            F_2 = int(strRead[0])
                            G_2 = int(strRead[1])
                            H_2 = float((self.handleInt2(strRead[3], strRead[2])) / 1000)
                            I_2 = float((self.handleInt2(strRead[5], strRead[4])) / 1000)
                            J_2 = float((self.handleInt2(strRead[7], strRead[6])) / 1000)
                            K_2 = float((self.handleInt2(strRead[9], strRead[8])) / 1000)
                            L_2 = float((self.handleInt2(strRead[11], strRead[10])) / 1000)

                            newItem_6 = QTableWidgetItem(str(H_2))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_6)

                            newItem_7 = QTableWidgetItem(str(I_2))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_7)

                            newItem_8 = QTableWidgetItem(str(J_2))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_8)

                            newItem_9 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_9)

                            newItem_10 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_10)

                            comboBoxList1 = [self._translate("Form", "系统操作"), self._translate("Form", "流程控制"),
                                             self._translate("Form", "输出口设置"), self._translate("Form", "回零运动"),
                                             self._translate("Form", "插补运动"), self._translate("Form", "独立运动")]
                            comboBoxList2 = {
                                self._translate("Form", "系统操作"): [self._translate("Form", "停止"),
                                                                  self._translate("Form", "启动"),
                                                                  self._translate("Form", "暂停"),
                                                                  self._translate("Form", "恢复"),
                                                                  self._translate("Form", "延时等待"),
                                                                  self._translate("Form", "等待电机完成"),
                                                                  self._translate("Form", "停止电机运动"),
                                                                  self._translate("Form", "常等待")],
                                self._translate("Form", "流程控制"): [self._translate("Form", "程序间跳转"),
                                                                  self._translate("Form", "程序循环"),
                                                                  self._translate("Form", "输入跳转"),
                                                                  self._translate("Form", "开启输入中断"),
                                                                  self._translate("Form", "关闭输入中断")],
                                self._translate("Form", "输出口设置"): [self._translate("Form", "输出口设置")],
                                self._translate("Form", "回零运动"): [self._translate("Form", "设置回零速度"),
                                                                  self._translate("Form", "启动回零")],
                                self._translate("Form", "插补运动"): [self._translate("Form", "设置点位速度"),
                                                                  self._translate("Form", "三轴相对运动"),
                                                                  self._translate("Form", "单轴绝对运动"),
                                                                  self._translate("Form", "XY绝对运动"),
                                                                  self._translate("Form", "XZ绝对运动"),
                                                                  self._translate("Form", "YZ绝对运动"),
                                                                  self._translate("Form", "三轴绝对运动"),
                                                                  self._translate("Form", "XY圆弧插补"),
                                                                  self._translate("Form", "XZ圆弧插补"),
                                                                  self._translate("Form", "YZ圆弧插补")],
                                self._translate("Form", "独立运动"): [self._translate("Form", "独立运动速度"),
                                                                  self._translate("Form", "相对运动"),
                                                                  self._translate("Form", "X绝对运动"),
                                                                  self._translate("Form", "Y绝对运动"),
                                                                  self._translate("Form", "Z绝对运动")]}

                            if F_2 == 1:
                                if G_2 == 1:
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“停止”
                                elif G_2 == 2:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 3:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 4:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“恢复”
                                elif G_2 == 5:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 6:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][5]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_2 = int(I_2)
                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_2 = int(J_2)
                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 7:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][6]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_2 = int(I_2)
                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_2 = int(J_2)
                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 8:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][7]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                            elif F_2 == 2:
                                if G_2 == 1:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 2:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_2 = int(I_2)
                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 3:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_2 = int(I_2)
                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_2 = int(J_2)
                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 4:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_2 = int(I_2)
                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_2 = int(J_2)
                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 5:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                            elif F_2 == 3:
                                if G_2 == 1:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "输出口设置")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_2 = int(I_2)
                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                            elif F_2 == 4:
                                if G_2 == 1:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "回零运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 2:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "回零运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_2 = int(I_2)
                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_2 = int(J_2)
                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                            elif F_2 == 5:
                                if G_2 == 1:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 2:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 3:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 4:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 5:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 6:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][5]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 7:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][6]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 8:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][7]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    L_2 = int(L_2)
                                    newItem_5 = QTableWidgetItem(str(L_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 9:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][8]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    L_2 = int(L_2)
                                    newItem_5 = QTableWidgetItem(str(L_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 10:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][9]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    L_2 = int(L_2)
                                    newItem_5 = QTableWidgetItem(str(L_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                            elif F_2 == 6:
                                if G_2 == 1:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_2 = int(H_2)
                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_2 == 2:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)
                                elif G_2 == 3:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_2 == 4:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_2 == 5:
                                    print()
                                    op = comboBoxList1[F_2 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_2))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                            if F_2 != 1 and F_2 != 2 and F_2 != 3 and F_2 != 4 and F_2 != 5 and F_2 != 6:
                                newItem_100 = QTableWidgetItem("无")
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_100)

                                newItem_101 = QTableWidgetItem("无")
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_101)

                                newItem_102 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_102)

                                newItem_103 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_103)

                                newItem_104 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_104)

                                newItem_105 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_105)

                                newItem_106 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_106)
                            j += 1
                        self.waitTime(100)
                        i += 1

                if D_1 != 0:
                    i = 1
                    while i <= C_1:
                        j = 1
                        while j <= 10:
                            read = master.execute(1, cst.HOLDING_REGISTERS, 0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                  12)
                            print(read)
                            strRead = [str(i) for i in read]
                            # strRead = [self.convertHex(i) for i in read]
                            print(strRead)

                            F_3 = int(strRead[0])
                            G_3 = int(strRead[1])
                            H_3 = float((self.handleInt2(strRead[3], strRead[2])) / 1000)
                            I_3 = float((self.handleInt2(strRead[5], strRead[4])) / 1000)
                            J_3 = float((self.handleInt2(strRead[7], strRead[6])) / 1000)
                            K_3 = float((self.handleInt2(strRead[9], strRead[8])) / 1000)
                            L_3 = float((self.handleInt2(strRead[11], strRead[10])) / 1000)

                            newItem_11 = QTableWidgetItem(str(H_3))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_11)

                            newItem_12 = QTableWidgetItem(str(I_3))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_12)

                            newItem_13 = QTableWidgetItem(str(J_3))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_13)

                            newItem_14 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_14)

                            newItem_15 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_15)

                            comboBoxList1 = [self._translate("Form", "系统操作"), self._translate("Form", "流程控制"),
                                             self._translate("Form", "输出口设置"), self._translate("Form", "回零运动"),
                                             self._translate("Form", "插补运动"), self._translate("Form", "独立运动")]
                            comboBoxList2 = {
                                self._translate("Form", "系统操作"): [self._translate("Form", "停止"),
                                                                  self._translate("Form", "启动"),
                                                                  self._translate("Form", "暂停"),
                                                                  self._translate("Form", "恢复"),
                                                                  self._translate("Form", "延时等待"),
                                                                  self._translate("Form", "等待电机完成"),
                                                                  self._translate("Form", "停止电机运动"),
                                                                  self._translate("Form", "常等待")],
                                self._translate("Form", "流程控制"): [self._translate("Form", "程序间跳转"),
                                                                  self._translate("Form", "程序循环"),
                                                                  self._translate("Form", "输入跳转"),
                                                                  self._translate("Form", "开启输入中断"),
                                                                  self._translate("Form", "关闭输入中断")],
                                self._translate("Form", "输出口设置"): [self._translate("Form", "输出口设置")],
                                self._translate("Form", "回零运动"): [self._translate("Form", "设置回零速度"),
                                                                  self._translate("Form", "启动回零")],
                                self._translate("Form", "插补运动"): [self._translate("Form", "设置点位速度"),
                                                                  self._translate("Form", "三轴相对运动"),
                                                                  self._translate("Form", "单轴绝对运动"),
                                                                  self._translate("Form", "XY绝对运动"),
                                                                  self._translate("Form", "XZ绝对运动"),
                                                                  self._translate("Form", "YZ绝对运动"),
                                                                  self._translate("Form", "三轴绝对运动"),
                                                                  self._translate("Form", "XY圆弧插补"),
                                                                  self._translate("Form", "XZ圆弧插补"),
                                                                  self._translate("Form", "YZ圆弧插补")],
                                self._translate("Form", "独立运动"): [self._translate("Form", "独立运动速度"),
                                                                  self._translate("Form", "相对运动"),
                                                                  self._translate("Form", "X绝对运动"),
                                                                  self._translate("Form", "Y绝对运动"),
                                                                  self._translate("Form", "Z绝对运动")]}

                            if F_3 == 1:
                                if G_3 == 1:
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“停止”
                                elif G_3 == 2:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 3:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)


                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“暂停”
                                elif G_3 == 4:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“恢复”
                                elif G_3 == 5:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 6:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][5]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_3 = int(I_3)
                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_3 = int(J_3)
                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 7:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][6]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_3 = int(I_3)
                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_3 = int(J_3)
                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 8:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][7]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                            elif F_3 == 2:
                                if G_3 == 1:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 2:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_3 = int(I_3)
                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 3:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_3 = int(I_3)
                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_3 = int(J_3)
                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 4:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_3 = int(I_3)
                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_3 = int(J_3)
                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 5:
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                            elif F_3 == 3:
                                if G_3 == 1:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "输出口设置")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_3 = int(I_3)
                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                            elif F_3 == 4:
                                if G_3 == 1:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "回零运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 2:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "回零运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    I_3 = int(I_3)
                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    J_3 = int(J_3)
                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                            elif F_3 == 5:
                                if G_3 == 1:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 2:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_4 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    newItem_5 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_3 == 3:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 4:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 5:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 6:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][5]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 7:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][6]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 8:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][7]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    L_3 = int(L_3)
                                    newItem_5 = QTableWidgetItem(str(L_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_3 == 9:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][8]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    L_3 = int(L_3)
                                    newItem_5 = QTableWidgetItem(str(L_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                                elif G_3 == 10:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][9]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)


                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_4)

                                    L_3 = int(L_3)
                                    newItem_5 = QTableWidgetItem(str(L_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_5)
                            elif F_3 == 6:
                                if G_3 == 1:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    H_3 = int(H_3)
                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 2:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)
                                elif G_3 == 3:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 4:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                                elif G_3 == 5:
                                    print()
                                    op = comboBoxList1[F_3 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_3))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_3)
                            if F_3 != 1 and F_3 != 2 and F_3 != 3 and F_3 != 4 and F_3 != 5 and F_3 != 6:
                                newItem_100 = QTableWidgetItem("无")
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 0, newItem_100)

                                newItem_101 = QTableWidgetItem("无")
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 1, newItem_101)

                                newItem_102 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 2, newItem_102)

                                newItem_103 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 3, newItem_103)

                                newItem_104 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 4, newItem_104)

                                newItem_105 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 5, newItem_105)

                                newItem_106 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * (i - 1) + (j - 1), 6, newItem_106)
                            j += 1
                        self.waitTime(100)
                        i += 1
                    else:
                        k = i - C_1
                        while k <= D_1:
                            read = master.execute(1, cst.HOLDING_REGISTERS, 0x0000 + 120 * C_1 + (k - 1)*12 , 12)
                            print(read)
                            strRead = [str(i) for i in read]
                            # strRead = [self.convertHex(i) for i in read]
                            print(strRead)

                            F_4 = int(strRead[0])
                            G_4 = int(strRead[1])
                            H_4 = float((self.handleInt2(strRead[3], strRead[2])) / 1000)
                            I_4 = float((self.handleInt2(strRead[5], strRead[4])) / 1000)
                            J_4 = float((self.handleInt2(strRead[7], strRead[6])) / 1000)
                            K_4 = float((self.handleInt2(strRead[9], strRead[8])) / 1000)
                            L_4 = float((self.handleInt2(strRead[11], strRead[10])) / 1000)

                            newItem_16 = QTableWidgetItem(str(H_4))
                            self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                            newItem_17 = QTableWidgetItem(str(I_4))
                            self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                            newItem_18 = QTableWidgetItem(str(J_4))
                            self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)

                            newItem_19 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(10 * C_1 + (k - 1), 5, newItem_19)

                            newItem_20 = QTableWidgetItem(str(0))
                            self.TableWidget.setItem(10 * C_1 + (k - 1), 6, newItem_20)

                            comboBoxList1 = [self._translate("Form", "系统操作"), self._translate("Form", "流程控制"),
                                             self._translate("Form", "输出口设置"), self._translate("Form", "回零运动"),
                                             self._translate("Form", "插补运动"), self._translate("Form", "独立运动")]
                            comboBoxList2 = {
                                self._translate("Form", "系统操作"): [self._translate("Form", "停止"),
                                                                  self._translate("Form", "启动"),
                                                                  self._translate("Form", "暂停"),
                                                                  self._translate("Form", "恢复"),
                                                                  self._translate("Form", "延时等待"),
                                                                  self._translate("Form", "等待电机完成"),
                                                                  self._translate("Form", "停止电机运动"),
                                                                  self._translate("Form", "常等待")],
                                self._translate("Form", "流程控制"): [self._translate("Form", "程序间跳转"),
                                                                  self._translate("Form", "程序循环"),
                                                                  self._translate("Form", "输入跳转"),
                                                                  self._translate("Form", "开启输入中断"),
                                                                  self._translate("Form", "关闭输入中断")],
                                self._translate("Form", "输出口设置"): [self._translate("Form", "输出口设置")],
                                self._translate("Form", "回零运动"): [self._translate("Form", "设置回零速度"),
                                                                  self._translate("Form", "启动回零")],
                                self._translate("Form", "插补运动"): [self._translate("Form", "设置点位速度"),
                                                                  self._translate("Form", "三轴相对运动"),
                                                                  self._translate("Form", "单轴绝对运动"),
                                                                  self._translate("Form", "XY绝对运动"),
                                                                  self._translate("Form", "XZ绝对运动"),
                                                                  self._translate("Form", "YZ绝对运动"),
                                                                  self._translate("Form", "三轴绝对运动"),
                                                                  self._translate("Form", "XY圆弧插补"),
                                                                  self._translate("Form", "XZ圆弧插补"),
                                                                  self._translate("Form", "YZ圆弧插补")],
                                self._translate("Form", "独立运动"): [self._translate("Form", "独立运动速度"),
                                                                  self._translate("Form", "相对运动"),
                                                                  self._translate("Form", "X绝对运动"),
                                                                  self._translate("Form", "Y绝对运动"),
                                                                  self._translate("Form", "Z绝对运动")]}

                            if F_4 == 1:
                                if G_4 == 1:
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_16 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“停止”
                                elif G_4 == 2:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_16 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“启动”
                                elif G_4 == 3:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_16 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“暂停”
                                elif G_4 == 4:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_16 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“恢复”
                                elif G_4 == 5:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                                elif G_4 == 6:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][5]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    I_4 = int(I_4)
                                    newItem_17 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    J_4 = int(J_4)
                                    newItem_18 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                                elif G_4 == 7:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][6]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    I_4 = int(I_4)
                                    newItem_17 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    J_4 = int(J_4)
                                    newItem_18 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                                elif G_4 == 8:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "系统操作")][7]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_16 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)

                                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                                    # 在指令的下拉框中设置“常等待”
                            elif F_4 == 2:
                                if G_4 == 1:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                                elif G_4 == 2:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    I_4 = int(I_4)
                                    newItem_17 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                                elif G_4 == 3:
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    I_4 = int(I_4)
                                    newItem_17 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    J_4 = int(J_4)
                                    newItem_18 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                                elif G_4 == 4:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    I_4 = int(I_4)
                                    newItem_17 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    J_4 = int(J_4)
                                    newItem_18 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                                elif G_4 == 5:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "流程控制")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_16 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_16)

                                    newItem_17 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_17)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                            elif F_4 == 3:
                                if G_4 == 1:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "输出口设置")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    I_4 = int(I_4)
                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_18 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_18)
                            elif F_4 == 4:
                                if G_4 == 1:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "回零运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 2:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "回零运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    I_4 = int(I_4)
                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    J_4 = int(J_4)
                                    newItem_3 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                            elif F_4 == 5:
                                if G_4 == 1:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 2:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)
                                elif G_4 == 3:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 4:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 5:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 6:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][5]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 7:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][6]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 8:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][7]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 5, newItem_4)

                                    L_4 = int(L_4)
                                    newItem_5 = QTableWidgetItem(str(L_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 6, newItem_5)
                                elif G_4 == 9:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][8]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 5, newItem_4)

                                    L_4 = int(L_4)
                                    newItem_5 = QTableWidgetItem(str(L_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 6, newItem_5)
                                elif G_4 == 10:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "插补运动")][9]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)

                                    newItem_4 = QTableWidgetItem(str(K_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 5, newItem_4)

                                    L_4 = int(L_4)
                                    newItem_5 = QTableWidgetItem(str(L_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 6, newItem_5)
                            elif F_4 == 6:
                                if G_4 == 1:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][0]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    H_4 = int(H_4)
                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(I_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(J_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 2:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][1]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)
                                elif G_4 == 3:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][2]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 4:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][3]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                                elif G_4 == 5:
                                    print()
                                    op = comboBoxList1[F_4 - 1]
                                    newItem_6 = QTableWidgetItem(str(op))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_6)

                                    ob = comboBoxList2[self._translate("Form", "独立运动")][4]
                                    newItem_7 = QTableWidgetItem(str(ob))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_7)

                                    newItem_1 = QTableWidgetItem(str(H_4))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_1)

                                    newItem_2 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_2)

                                    newItem_3 = QTableWidgetItem(str(0))
                                    self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_3)
                            if F_4 != 1 and F_4 != 2 and F_4 != 3 and F_4 != 4 and F_4 != 5 and F_4 != 6:
                                newItem_100 = QTableWidgetItem("无")
                                self.TableWidget.setItem(10 * C_1 + (k - 1), 0, newItem_100)

                                newItem_101 = QTableWidgetItem("无")
                                self.TableWidget.setItem(10 * C_1 + (k - 1), 1, newItem_101)

                                newItem_102 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * C_1 + (k - 1), 2, newItem_102)

                                newItem_103 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * C_1 + (k - 1), 3, newItem_103)

                                newItem_104 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * C_1 + (k - 1), 4, newItem_104)

                                newItem_105 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * C_1 + (k - 1), 5, newItem_105)

                                newItem_106 = QTableWidgetItem(str(0))
                                self.TableWidget.setItem(10 * C_1 + (k - 1), 6, newItem_106)
                            k += 1

    def hello1(self):
        for i in range(5):
            s = ''
            ss = self.TableWidget.item(0, i).text()
            print('s3__send_text=' + ss)
            if ss.find('请输入要发送数据') >= -1:
                self.TableWidget.item(0, i).setText('')
            ss = self.TableWidget.item(0, i).text()
            if ss == '':
                self.TableWidget.item(0, i).setText(s)
            else:
                self.TableWidget.item(0, i).setText(ss + ' ' + s)
            i += 1

    def on_btnAutoHeight_clicked(self):
        """
        列的宽度随着内容变化
        :return:
        """
        self.TableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.TableWidget.setColumnWidth(0,95)
        self.TableWidget.setColumnWidth(1,95)
        self.TableWidget.setColumnWidth(2,65)
        self.TableWidget.setColumnWidth(3,65)
        self.TableWidget.setColumnWidth(4,65)
        self.TableWidget.setColumnWidth(5,65)
        self.TableWidget.setColumnWidth(6,65)
        self.TableWidget.horizontalHeader().setStyleSheet("QHeaderView::section{background:rgb(240,180,140);}")
        self.TableWidget.setAlternatingRowColors(True)

    def ON_SetUp_Port(self):
        Dig = SerialPortDialogFunc()
        Dig.setport(self.port)
        Dig.setbotelv(self.botelv)
        Dig.setshujuwei(self.shujuwei)
        Dig.setjiaoyanwei(self.jiaoyanwei)
        Dig.setstop(self.stop)
        vle = Dig.exec()
        if vle == QDialog.Accepted:
            self.port = Dig.getport()
            self.botelv = Dig.getbotelv()
            self.shujuwei = Dig.getshujuwei()
            self.jiaoyanwei = Dig.getjiaoyanwei()
            self.stop = Dig.getstop()
        return vle

    def retranslateTableWidgetUi(self):
        row = self.TableWidget.rowCount()
        for rowindex in range(row):
            col1 = self.TableWidget.item(rowindex,0)
            col2 = self.TableWidget.item(rowindex,1)
            if col1 != None and col2 != None:
                col1text = col1.text()
                col2text = col2.text()
                col1map = [self._translate("Form", "系统操作"), self._translate("Form", "流程控制"),
                                 self._translate("Form", "输出口设置"), self._translate("Form", "回零运动"),
                                 self._translate("Form", "插补运动"), self._translate("Form", "独立运动")]
                col2map= {
                    self._translate("Form", "系统操作"): [self._translate("Form", "停止"), self._translate("Form", "启动"),
                                                      self._translate("Form", "暂停"), self._translate("Form", "恢复"),
                                                      self._translate("Form", "延时等待"),
                                                      self._translate("Form", "等待电机完成"),
                                                      self._translate("Form", "停止电机运动"),
                                                      self._translate("Form", "常等待")],
                    self._translate("Form", "流程控制"): [self._translate("Form", "程序间跳转"),
                                                      self._translate("Form", "程序循环"),
                                                      self._translate("Form", "输入跳转"),
                                                      self._translate("Form", "开启输入中断"),
                                                      self._translate("Form", "关闭输入中断")],
                    self._translate("Form", "输出口设置"): [self._translate("Form", "输出口设置")],
                    self._translate("Form", "回零运动"): [self._translate("Form", "设置回零速度"),
                                                      self._translate("Form", "启动回零")],
                    self._translate("Form", "插补运动"): [self._translate("Form", "设置点位速度"),
                                                      self._translate("Form", "三轴相对运动"),
                                                      self._translate("Form", "单轴绝对运动"),
                                                      self._translate("Form", "XY绝对运动"),
                                                      self._translate("Form", "XZ绝对运动"),
                                                      self._translate("Form", "YZ绝对运动"),
                                                      self._translate("Form", "三轴绝对运动"),
                                                      self._translate("Form", "XY圆弧插补"),
                                                      self._translate("Form", "XZ圆弧插补"),
                                                      self._translate("Form", "YZ圆弧插补")],
                    self._translate("Form", "独立运动"): [self._translate("Form", "独立运动速度"),
                                                      self._translate("Form", "相对运动"),
                                                      self._translate("Form", "X绝对运动"),
                                                      self._translate("Form", "Y绝对运动"),
                                                      self._translate("Form", "Z绝对运动")]}
                if self.Language == "ENGLISH":
                    """
                  中文到英文
                  """
                    col1.setText(self._translate("Form",col1text))
                    col2.setText(self._translate("Form",col2text))
                if self.Language == "CHINESE":
                    # todo python的键值对使用的方法
                    col1.setText(self.retranslatexlxs(col1text))
                    col2.setText(self.retranslatexlxs(col2text))

    def retranslatexlxs(self,value) -> str:
        # 返回的是字符串
        xlsxfilename = os.path.join(os.path.dirname(__file__),"中英文对照.xlsx")
        data = xlrd.open_workbook(xlsxfilename)
        table = data.sheets()[0]
        A_1 = table.nrows  # 获取行数  （A_1肯定是偶数行）14
        key = ""
        i = 0
        while i < A_1:
            command_info_list_i = table.row_values(i)  # 获取sheet中第i行的数据
            print(command_info_list_i)
            a = command_info_list_i[0]
            b = command_info_list_i[2]
            if b == value:
                key = a
                break
            i += 1
        return key

    def retranslateUi(self, Form):
        super().retranslateUi(Form)     # todo python的类的重载多态
        self.retranslateTableWidgetUi()
        if self.ser != None:
            if self.ser.isOpen():
                self.s1__lb_42.setText(self._translate("Form", "当前状态：已连接控制器"))
                # self.s1__lb_43.setText(self._translate("Form", "已连接"))
            else:
                self.s1__lb_42.setText(self._translate("Form", "当前状态：已断开控制器"))
                # self.s1__lb_43.setText(self._translate("Form", "已断开"))
        else:
            self.s1__lb_42.setText("")
            # self.s1__lb_43.setText("")

    def port_open(self):

        global master
        while True:
            print(self.port)
            if self.port != "":
                # self.ser.port = self.port
                try:
                    master = modbus_rtu.RtuMaster(serial.Serial(port=self.port,
                     baudrate=int(self.botelv), bytesize=int(self.shujuwei), parity=self.jiaoyanwei, stopbits=float(self.stop)))
                    master.set_timeout(2, True)
                    self.ser = master._serial

                    self.s1__lb_42.setText(self._translate("Form", "当前状态：已连接控制器"))
                    # self.s1__lb_42.setFont(QFont("华文楷书", 10.5))
                    self.s1__lb_42.setFont(QFont("华文楷书", round(10.5)))
                    self.s1__lb_42.setStyleSheet("color:rgb(250,0,0);")

                    self.readini.set("common", "port",self.port)
                    self.readini.set("common", "botelv", self.botelv)
                    self.readini.set("common", "shujuwei", self.shujuwei)
                    self.readini.set("common", "jiaoyanwei", self.jiaoyanwei)
                    self.readini.set("common", "stop", self.stop)
                    self.readini.write(open(self.conf_path, "w"))

                except Exception as err:
                    QMessageBox.critical(self, "错误提示" , "此串口不能被打开！" + str(err))

                if self.ser is not None and self.ser.isOpen():
                    self.open_button.setEnabled(False)
                    break
                else:
                    vle = self.ON_SetUp_Port()
                    print(vle)
                    if vle != QDialog.Accepted:
                       break

            else:
                vle = self.ON_SetUp_Port()
                print(vle)
                if vle != QDialog.Accepted:
                    break

    def port_close(self):
        if self.timer1 > 0:
            self.killTimer(self.timer1)
            self.timer1 = 0

        try:
            self.ser.close()
            self.s1__lb_42.setText(self._translate("Form", "当前状态：已断开控制器"))
            # self.s1__lb_43.setText(self._translate("Form", "已断开"))
        except:
            pass
        self.open_button.setEnabled(True)

#  定义发送数据
    def convertHex(self, n):
        sHex = hex(n)[2:]
        return sHex if len(sHex) > 1 else '0' + sHex

    def data_send(self):
        """
        “下载程序”按钮执行的动作
        :return:
        """
        # global a
        # global b
        # global c
        # global d
        Func_code_READ_COILS = 0x01  # 读线圈
        Func_code_InputStatus = 0x02
        Func_code_READ_HOLDING_REGISTERS = 0x03  # 读寄存器状态
        Func_code_WRITE_SINGLE_COIL = 0x05  # 写单线圈
        Func_code_WRITE_SINGLE_REGISTER = 0x06  # 写单寄存器
        Func_code_WRITE_MULTIPLE_COILS = 0x0F  # 写多线圈
        Func_code_WRITE_MULTIPLE_REGISTERS = 0x10  # 写多寄存器
        # self.s1__box_9.currentText().strip()
        logger = modbus_tk.utils.create_logger("console")
        logger.info("connected")
        A_1 = int(self.lineEdit_28.text())  # 表示获取文本框中要下载的行数
        print(A_1)

        B_1 = A_1 * 12  # 每行12个寄存器 B_1 = A_1 * 12 表示要下载填写数据的寄存器总数
        print(B_1)

        C_1 = A_1 // 10  # C_1表示的整10行数
        print(C_1)

        D_1 = A_1 % 10  # 行数的个位数
        print(D_1)

        E_1 = C_1 + 1  # E_1是下载的次数
        print(E_1)

        if A_1 == 0:
            QMessageBox.critical(self, "Port Error\n", "没有工程参数！")
            return None

        for i in range(1):
            logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x219C, output_value=[A_1]))

        if A_1 != 0:
            if E_1 == 1:
                i = 1
                while i <= D_1:
                    a = self.TableWidget.item(i - 1, 0).text()
                    b = self.TableWidget.item(i - 1, 1).text()

                    a_1 = float(self.TableWidget.item(i - 1, 2).text())  # 对“参数1”取值
                    print(a_1)
                    b_1 = a_1 * 1000
                    print(b_1)
                    c_1 = int(b_1 % 65536)
                    if c_1 > 32767:
                        c_1 = int(c_1 - 65536)
                        print(c_1)
                    if 0 < c_1 < 32767:
                        c_1 = int(c_1)
                        print(c_1)
                    d_1 = int(b_1 // 65536)
                    print(d_1)

                    a_2 = float(self.TableWidget.item(i - 1, 3).text())  # 对“参数2”取值
                    print(a_2)
                    b_2 = a_2 * 1000
                    print(b_2)
                    c_2 = int(b_2 % 65536)
                    if c_2 > 32767:
                        c_2 = int(c_2 - 65536)
                        print(c_2)
                    if 0 < c_2 < 32767:
                        c_2 = int(c_2)
                        print(c_2)
                    d_2 = int(b_2 // 65536)
                    print(d_2)

                    a_3 = float(self.TableWidget.item(i - 1, 4).text())  # 对“参数2”取值
                    print(a_3)
                    b_3 = a_3 * 1000
                    print(b_3)
                    c_3 = int(b_3 % 65536)
                    if c_3 > 32767:
                        c_3 = int(c_3 - 65536)
                        print(c_3)
                    if 0 < c_3 < 32767:
                        c_3 = int(c_3)
                        print(c_3)
                    d_3 = int(b_3 // 65536)
                    print(d_3)

                    a_4 = float(self.TableWidget.item(i - 1, 5).text())  # 对“参数4”取值
                    print(a_4)
                    b_4 = a_4 * 1000
                    print(b_4)
                    c_4 = int(b_4 % 65536)
                    if c_4 > 32767:
                        c_4 = int(c_4 - 65536)
                        print(c_4)
                    if 0 < c_4 < 32767:
                        c_4 = int(c_4)
                        print(c_4)
                    d_4 = int(b_4 // 65536)
                    print(d_4)

                    a_5 = float(self.TableWidget.item(i - 1, 6).text())  # 对“参数5”取值
                    print(a_5)
                    b_5 = a_5 * 1000
                    print(b_5)
                    c_5 = int(b_5 % 65536)
                    if c_5 > 32767:
                        c_5 = int(c_5 - 65536)
                        print(c_5)
                    if 0 < c_5 < 32767:
                        c_5 = int(c_5)
                        print(c_5)
                    d_5 = int(b_5 // 65536)
                    print(d_5)

                    if a == "系统操作" and b == "停止":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 1, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "系统操作" and b == "启动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 2, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "系统操作" and b == "暂停":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 3, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "系统操作" and b == "恢复":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 4, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "系统操作" and b == "延时等待":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 5, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "系统操作" and b == "等待电机完成":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 6, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "系统操作" and b == "停止电机运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 7, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "系统操作" and b == "常等待":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[1, 8, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "流程控制" and b == "程序间跳转":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[2, 1, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "流程控制" and b == "程序循环":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[2, 2, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "流程控制" and b == "输入跳转":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[2, 3, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "流程控制" and b == "开启输入中断":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[2, 4, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "流程控制" and b == "关闭输入中断":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[2, 5, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "输出口设置" and b == "输出口设置":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[3, 1, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "回零运动" and b == "设置回零速度":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[4, 1, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "回零运动" and b == "启动回零":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[4, 2, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "设置点位速度":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 1, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "三轴相对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 2, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "单轴绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 3, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "XY绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 4, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "XZ绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 5, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "YZ绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 6, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "三轴绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 7, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "XY圆弧插补":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 8, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "XZ圆弧插补":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 9, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "插补运动" and b == "YZ圆弧插补":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[5, 10, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "独立运动" and b == "独立运动速度":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[6, 1, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "独立运动" and b == "相对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[6, 2, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "独立运动" and b == "X绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[6, 3, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "独立运动" and b == "Y绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[6, 4, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))

                    if a == "独立运动" and b == "Z绝对运动":
                        logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 + 12 * (i - 1),
                                                   output_value=[6, 5, c_1, d_1, c_2, d_2, c_3, d_3, c_4, d_4, c_5,
                                                                 d_5]))
                    self.waitTime(50)
                    i += 1

            if E_1 >= 2:
                if D_1 == 0:
                    i = 1
                    while i <= C_1:
                        j = 1
                        while j <= 10:

                            a = self.TableWidget.item(10 * (i - 1) + (j - 1), 0).text()
                            b = self.TableWidget.item(10 * (i - 1) + (j - 1), 1).text()

                            a_6 = float(self.TableWidget.item(10 * (i - 1) + (j - 1), 2).text())  # 对“参数1”取值
                            print(a_6)
                            b_6 = a_6 * 1000
                            print(b_6)
                            c_6 = int(b_6 % 65536)
                            if c_6 > 32767:
                                c_6 = int(c_6 - 65536)
                                print(c_6)
                            if 0 < c_6 < 32767:
                                c_6 = int(c_6)
                                print(c_6)
                            d_6 = int(b_6 // 65536)
                            print(d_6)

                            a_7 = float(self.TableWidget.item(10 * (i - 1) + (j - 1), 3).text())  # 对“参数1”取值
                            print(a_7)
                            b_7 = a_7 * 1000
                            print(b_7)
                            c_7 = int(b_7 % 65536)
                            if c_7 > 32767:
                                c_7 = int(c_7 - 65536)
                                print(c_7)
                            if 0 < c_7 < 32767:
                                c_7 = int(c_7)
                                print(c_7)
                            d_7 = int(b_7 // 65536)
                            print(d_7)

                            a_8 = float(self.TableWidget.item(10 * (i - 1) + (j - 1), 4).text())  # 对“参数1”取值
                            print(a_8)
                            b_8 = a_8 * 1000
                            print(b_8)
                            c_8 = int(b_8 % 65536)
                            if c_8 > 32767:
                                c_8 = int(c_8 - 65536)
                                print(c_8)
                            if 0 < c_8 < 32767:
                                c_8 = int(c_8)
                                print(c_8)
                            d_8 = int(b_8 // 65536)
                            print(d_8)

                            a_9 = float(self.TableWidget.item(10 * (i - 1) + (j - 1), 5).text())  # 对“参数1”取值
                            print(a_9)
                            b_9 = a_9 * 1000
                            print(b_9)
                            c_9 = int(b_9 % 65536)
                            if c_9 > 32767:
                                c_9 = int(c_9 - 65536)
                                print(c_9)
                            if 0 < c_9 < 32767:
                                c_9 = int(c_9)
                                print(c_9)
                            d_9 = int(b_9 // 65536)
                            print(d_9)

                            a_10 = float(self.TableWidget.item(10 * (i - 1) + (j - 1), 6).text())  # 对“参数1”取值
                            print(a_10)
                            b_10 = a_10 * 1000
                            print(b_10)
                            c_10 = int(b_10 % 65536)
                            if c_10 > 32767:
                                c_10 = int(c_10 - 65536)
                                print(c_10)
                            if 0 < c_10 < 32767:
                                c_10 = int(c_10)
                                print(c_10)
                            d_10 = int(b_10 // 65536)
                            print(d_10)

                            if a == "无":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]))

                            if a == "系统操作" and b == "停止":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[1, 1, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "系统操作" and b == "启动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[1, 2, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "系统操作" and b == "暂停":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[1, 3, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "系统操作" and b == "恢复":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[1, 4, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "系统操作" and b == "延时等待":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[1, 5, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "系统操作" and b == "等待电机完成":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[1, 6, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "系统操作" and b == "停止电机运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[1, 7, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "系统操作" and b == "常等待":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[1, 8, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "流程控制" and b == "程序间跳转":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[2, 1, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "流程控制" and b == "程序循环":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[2, 2, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "流程控制" and b == "输入跳转":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[2, 3, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "流程控制" and b == "开启输入中断":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[2, 4, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "流程控制" and b == "关闭输入中断":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[2, 5, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "输出口设置" and b == "输出口设置":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[3, 1, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "回零运动" and b == "设置回零速度":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[4, 1, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "回零运动" and b == "启动回零":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[4, 2, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "设置点位速度":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[5, 1, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "三轴相对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[5, 2, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "单轴绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[5, 3, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "XY绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[5, 4, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "XZ绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[5, 5, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "YZ绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[5, 6, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "三轴绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[5, 7, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "XY圆弧插补":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12 ,
                                                           output_value=[5, 8, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "XZ圆弧插补":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[5, 9, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "插补运动" and b == "YZ圆弧插补":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[5, 10, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "独立运动" and b == "独立运动速度":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[6, 1, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "独立运动" and b == "相对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[6, 2, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "独立运动" and b == "X绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[6, 3, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "独立运动" and b == "Y绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[6, 4, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))

                            if a == "独立运动" and b == "Z绝对运动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + (i - 1) * 120 + (j - 1) * 12,
                                                           output_value=[6, 5, c_6, d_6, c_7, d_7, c_8, d_8, c_9, d_9,
                                                                         c_10, d_10]))
                            self.waitTime(200)
                            j += 1
                        i += 1

                if D_1 != 0:
                    i = 1
                    while i <= C_1:
                        j = 1
                        while j <= 10:
                            logger.info(
                                master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                               0x0000 + 120 * (i - 1) + 12 * (j - 1),
                                               output_value=[1, 6, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]))
                            j += 1
                        i += 1
                    else:
                        k = 1
                        while k <= D_1:
                            a = self.TableWidget.item(10 * C_1 + (k - 1), 0).text()
                            b = self.TableWidget.item(10 * C_1 + (k - 1), 1).text()

                            a_16 = float(self.TableWidget.item(10 * C_1 + (k - 1), 2).text())  # 对“参数1”取值
                            print(a_16)
                            b_16 = a_16 * 1000
                            print(b_16)
                            c_16 = int(b_16 % 65536)
                            if c_16 > 32767:
                                c_16 = int(c_16 - 65536)
                                print(c_16)
                            if 0 < c_16 < 32767:
                                c_16 = int(c_16)
                                print(c_16)
                            d_16 = int(b_16 // 65536)
                            print(d_16)

                            a_17 = float(self.TableWidget.item(10 * C_1 + (k - 1), 3).text())  # 对“参数1”取值
                            print(a_17)
                            b_17 = a_17 * 1000
                            print(b_17)
                            c_17 = int(b_17 % 65536)
                            if c_17 > 32767:
                                c_17 = int(c_17 - 65536)
                                print(c_17)
                            if 0 < c_17 < 32767:
                                c_17 = int(c_17)
                                print(c_17)
                            d_17 = int(b_17 // 65536)
                            print(d_17)

                            a_18 = float(self.TableWidget.item(10 * C_1 + (k - 1),4).text())  # 对“参数1”取值
                            print(a_18)
                            b_18 = a_18 * 1000
                            print(b_18)
                            c_18 = int(b_18 % 65536)
                            if c_18 > 32767:
                                c_18 = int(c_18 - 65536)
                                print(c_18)
                            if 0 < c_18 < 32767:
                                c_18 = int(c_18)
                                print(c_18)
                            d_18 = int(b_18 // 65536)
                            print(d_18)

                            a_19 = float(self.TableWidget.item(10 * C_1 + (k - 1), 5).text())  # 对“参数1”取值
                            print(a_19)
                            b_19 = a_19 * 1000
                            print(b_19)
                            c_19 = int(b_19 % 65536)
                            if c_19 > 32767:
                                c_19 = int(c_19 - 65536)
                                print(c_19)
                            if 0 < c_19 < 32767:
                                c_19 = int(c_19)
                                print(c_19)
                            d_19 = int(b_19 // 65536)
                            print(d_19)

                            a_20 = float(self.TableWidget.item(10 * C_1 + (k - 1), 6).text())  # 对“参数1”取值
                            print(a_20)
                            b_20 = a_20 * 1000
                            print(b_20)
                            c_20 = int(b_20 % 65536)
                            if c_20 > 32767:
                                c_20 = int(c_20 - 65536)
                                print(c_20)
                            if 0 < c_20 < 32767:
                                c_20 = int(c_20)
                                print(c_20)
                            d_20 = int(b_20 // 65536)
                            print(d_20)

                            if a == "无":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]))

                            if a == "系统操作" and b == "停止":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                   0x0000 +120 * C_1 + 12 * (k - 1) ,
                                                   output_value=[1, 1, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "系统操作" and b == "启动":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[1, 2, c_16, d_16, c_17, d_17, c_18, d_18, c_19,
                                                                         d_19, c_20, d_20]))

                            if a == "系统操作" and b == "暂停":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[1, 3, c_16, d_16, c_17, d_17, c_18, d_18, c_19,
                                                                         d_19, c_20, d_20]))

                            if a == "系统操作" and b == "恢复":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[1, 4, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "系统操作" and b == "延时等待":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[1, 5, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "系统操作" and b == "等待电机完成":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[1, 6, c_16, d_16, c_17, d_17, c_18, d_18, c_19,
                                                                         d_19, c_20, d_20]))

                            if a == "系统操作" and b == "停止电机运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[1, 7, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "系统操作" and b == "常等待":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[1, 8, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "流程控制" and b == "程序间跳转":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[2, 1, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "流程控制" and b == "程序循环":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[2, 2, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "流程控制" and b == "输入跳转":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[2, 3, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "流程控制" and b == "开启输入中断":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[2, 4, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "流程控制" and b == "关闭输入中断":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[2, 5, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "输出口设置" and b == "输出口设置":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[3, 1, c_16, d_16, c_17, d_17, c_18, d_18, c_19,
                                                                         d_19, c_20, d_20]))

                            if a == "回零运动" and b == "设置回零速度":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[4, 1, c_16, d_16, c_17, d_17, c_18, d_18, c_19,
                                                                         d_19, c_20, d_20]))

                            if a == "回零运动" and b == "启动回零":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[4, 2, c_16, d_16, c_17, d_17, c_18, d_18, c_19,
                                                                         d_19, c_20, d_20]))

                            if a == "插补运动" and b == "设置点位速度":
                                logger.info(master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,
                                                           0x0000 + 120 * C_1 + 12 * (k - 1),
                                                           output_value=[5, 1, c_16, d_16, c_17, d_17, c_18, d_18, c_19,
                                                                         d_19, c_20, d_20]))

                            if a == "插补运动" and b == "三轴相对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 2, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "单轴绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 3, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "XY绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 4, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "XZ绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 5, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "YZ绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 6, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "三轴绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS,0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 7, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "XY圆弧插补":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 8, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "XZ圆弧插补":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 9, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "插补运动" and b == "YZ圆弧插补":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[5, 10, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "独立运动" and b == "独立运动速度":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[6, 1, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "独立运动" and b == "相对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[6, 2, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "独立运动" and b == "X绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[6, 3, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "独立运动" and b == "Y绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[6, 4, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))

                            if a == "独立运动" and b == "Z绝对运动":
                                logger.info(
                                    master.execute(1, cst.WRITE_MULTIPLE_REGISTERS, 0x0000 +120 * C_1 + 12 * (k - 1),
                                                   output_value=[6, 5, c_16, d_16, c_17, d_17, c_18, d_18, c_19, d_19,
                                                                 c_20, d_20]))
                            self.waitTime(200)
                            k += 1

    # todo 定义接收数据
    def XXXXXXXXXXXXXXdata_receive(self):
        try:
            num = self.ser.inWaiting()
        except:
            self.port_close()
            return None
        if num > 0:
            data = self.ser.read(num)
            num = len(data)
            # hex显示

            if self.hex_receive.checkState():
                out_s = ''
                for i in range(0, len(data)):
                    out_s = out_s + '{:02X}'.format(data[i]) + ' '
                    self.TableWidget.item(1, 0).insertPlainText(out_s)
            else:
                # 串口接收到的字符串为b'123',要转化成unicode字符串才能输出到窗口中去
                self.TableWidget.item(1, 0).setText(str(data))

            # 统计接收字符的数量
            self.data_num_received += num
            self.lineEdit.setText(str(self.data_num_received))

            # 获取到text光标
            # textCursor = self.TableWidget.item(1,0).textCursor()
            # 滚动到底部
            # textCursor.movePosition(textCursor.End)
            # 设置光标到text中去
            # self.TableWidget.item(1,0).setTextCursor(textCursor)
        else:
            pass

    # 定时发送数据
    def data_send_timer(self):
        if self.timer_send_cb.isChecked():
            self.timer_send.start(int(self.lineEdit_3.text()))
            self.lineEdit_3.setEnabled(False)
        else:
            self.timer_send.stop()
            self.lineEdit_3.setEnabled(True)

    # 要在这边添加的是CRC校验的的功能，在ui文件添加
    def nor_com1(self,V1, V2, V3, V4, V5, V6, CRC1, CRC2) -> bool:
        modbus_data = [V1, V2, V3, V4, V5, V6]
        # modbus_data = [hex(V1),hex(V2),hex(V3),hex(V4),hex(V5),hex(V6),hex(V7)]
        crc = modbus_crc(modbus_data)
        print(crc)
        isok = CRC2 == crc[0] and CRC1 == crc[1]
        return isok

    ####################################################################################
    def onChange(self):
        global num1
        global num2
        # X轴
        num1 = ''  # 导程里边填写的数据
        num2 = ''  # 细分数里边填写的数据
        if self.lineEdit_13.text() != '':
            num1 = float(self.lineEdit_13.text())
        if self.lineEdit_16.text() != '':
            num2 = float(self.lineEdit_16.text())
        print(num1)
        if num1 != '' and num2 != '':
            v = float(num1) / float(num2)
            v = float('%.3f' % v)
            self.lineEdit_10.setText(str(v))
        # Y轴
        num3 = ''
        num4 = ''
        if self.lineEdit_14.text() != '':
            num3 = float(self.lineEdit_14.text())
        if self.lineEdit_17.text() != '':
            num4 = float(self.lineEdit_17.text())
        if num3 != '' and num4 != '':
            v = float(num3) / float(num4)
            v = float('%.3f' % v)
            self.lineEdit_11.setText(str(v))
        # z轴
        num5 = ''
        num6 = ''
        if self.lineEdit_15.text() != '':
            num5 = float(self.lineEdit_15.text())
        if self.lineEdit_18.text() != '':
            num6 = float(self.lineEdit_18.text())
        if num5 != '' and num6 != '':
            v = float(num5) / float(num6)
            v = float('%.3f' % v)
            self.lineEdit_12.setText(str(v))

    ####################################################################################
    # 计算脉冲当量那个位置的
    def moveup(self):
        # self.TableWidget
        row = self.TableWidget.currentIndex().row()  # 获取行号
        if row == 0:
            QMessageBox.critical(self, "Error", "第一行不能上移！")
            return
        column = self.TableWidget.currentIndex().column()  # 获取列序
        # todo 获取商品id
        # urlsmall = str(self.TableWidget.model().dataAt(r,2))
        contents = self.TableWidget.selectedItems()[0].text()  # 获取选中文本内容
        print("选择的内容为：", contents)
        print("所选的内容所在的行为：", row)
        print("所选的内容所在的列为：", column)

        row2 = self.TableWidget.currentIndex().row() - 1  # 获取行号
        print('row2=', row2)
        # column = self.TableWidget.currentIndex().column()  # 获取列序
        # todo 获取商品id
        # urlsmall = str(self.TableWidget.model().dataAt(r,2))
        row2item = self.TableWidget.item(row2, 0)
        if row2item == None:
            print("上一个选择的内容为：none")
        else:
            contents2 = self.TableWidget.item(row2, 0).text()  # 获取选中文本内容
            print("上一个选择的内容为：", contents2)
            for i in range(self.TableWidget.columnCount()):
                # todo 上下行互换内容 temp保存中间变量
                temp = ''
                contents = self.TableWidget.item(row, i)
                contents2 = self.TableWidget.item(row2, i)
                if contents != None:
                    contents = contents.text()
                    temp = contents2.text()
                    print('temp-', temp)
                    newItem = QTableWidgetItem(contents)
                    self.TableWidget.setItem(row2, i, newItem)

                if contents2 != None:
                    newItem2 = QTableWidgetItem(temp)
                    self.TableWidget.setItem(row, i, newItem2)

            self.TableWidget.setCurrentCell(row2, 0)  # 跳到上一行来
            # self.TableWidget.setItem()

    def movedown(self):
        row = self.TableWidget.currentIndex().row()  # 获取行号
        column = self.TableWidget.currentIndex().column()  # 获取列序
        # todo 获取商品id
        contents = self.TableWidget.selectedItems()[0].text()  # 获取选中文本内容
        print("选择的内容为：", contents)
        print("所选的内容所在的行为：", row)
        print("所选的内容所在的列为：", column)

        row2 = self.TableWidget.currentIndex().row() + 1  # 获取行号
        print('row2=', row2)
        # column = self.TableWidget.currentIndex().column()  # 获取列序
        # todo 获取商品id
        # urlsmall = str(self.TableWidget.model().dataAt(r,2))
        row2item = self.TableWidget.item(row2, 0)
        if row2item == None:
            print("上一个选择的内容为：none")
        else:
            contents2 = self.TableWidget.item(row2, 0).text()  # 获取选中文本内容
            print("上一个选择的内容为：", contents2)
            for i in range(self.TableWidget.columnCount()):
                # todo 上下行互换内容 temp保存中间变量
                temp = ''
                contents = self.TableWidget.item(row, i)
                contents2 = self.TableWidget.item(row2, i)
                if contents != None:
                    contents = contents.text()
                    temp = contents2.text()
                    print('temp-', temp)
                    newItem = QTableWidgetItem(contents)
                    self.TableWidget.setItem(row2, i, newItem)

                if contents2 != None:
                    newItem2 = QTableWidgetItem(temp)
                    self.TableWidget.setItem(row, i, newItem2)

            self.TableWidget.setCurrentCell(row2, 0)  # 跳到上一行来

    # todo self.tableWidget.AddRow(col)#添加指定行
    def AddRow(self):
        a = int(self.lineEdit_28.text())
        b = a + 1
        self.lineEdit_28.setText(str(b))
        row = self.TableWidget.currentIndex().row()  # 获取行号
        self.TableWidget.insertRow(row+1)
        newItem = QTableWidgetItem('  ---请选择---')
        self.TableWidget.setItem(row+1, 0, newItem)
        newItem2 = QTableWidgetItem('  ---请选择---')
        self.TableWidget.setItem(row+1, 1, newItem2)

    def handleInt2(self,b,a):
        b = self.handleInt(int(b))  # 高位寄存器
        a = int(a)  # 低位寄存器
        vle = b * 65536 + a
        return vle

    def handleInt(self,v):
        isNegative= False
        vle = v
        if vle & 0x8000 != 0:
            isNegative = True
        if isNegative:
            vle = vle & 0x7fff
            vle = ~vle
            vle = vle & 0x7fff
            vle = vle + 1
            vle = 0 - vle

        return vle

    # todo self.tableWidget.removeRow(col)#删除指定行
    def DeleteRow(self):
        row = self.TableWidget.currentIndex().row()  # 获取行号
        print(row)
        a = int(self.lineEdit_28.text())
        B = row+1
        if B != a:
            self.TableWidget.removeRow(row)
            A = self.TableWidget.item(row,0).text()
            if A == '':
                print("删除就行了")

            else:
                a = int(self.lineEdit_28.text())
                b = a - 1
                self.lineEdit_28.setText(str(b))
        else:
            self.TableWidget.removeRow(row)
            a = int(self.lineEdit_28.text())
            b = a - 1
            self.lineEdit_28.setText(str(b))
            print("闭嘴吧")

    # todo self.tableWidget.CopyRow(col)#复制指定行
    def CopyRow(self):
        row = self.TableWidget.currentIndex().row()  # 获取行号
        # todo 重置粘贴板
        self.cupboard = []
        for i in range(self.TableWidget.columnCount()):
            content = self.TableWidget.item(row, i)
            if content != None:
                self.cupboard.append(content.text())
            else:
                self.cupboard.append('')

    # todo self.tableWidget.PasteRow(col)#粘贴指定行
    def PasteRow(self):
        row = self.TableWidget.currentIndex().row()  # 获取行号
        row = row+1
        print(row)
        if self.cupboard.__len__() == 0:
            print('粘贴板无内容')
            return
        else:
            print('粘贴')
            for i in range(self.TableWidget.columnCount()):
                print('粘贴', self.cupboard[i])
                newItem = QTableWidgetItem(self.cupboard[i])
                self.TableWidget.setItem(row-1, i, newItem)
        a = int(self.lineEdit_28.text())
        if row <= a:
            print("没事了")
        else:
            a = int(self.lineEdit_28.text())
            b=a+1
            self.lineEdit_28.setText(str(b))

    # todo 将表格中的数据内容保存到excel表中（3.11晚完成）
    def saveExcelfiles(self):
        s2fname = QFileDialog.getSaveFileName(self,'save file',self.ProjectName ,"xlsx files (*.xlsx);;all files(*.*)")[0]
        if os.path.exists(s2fname):
            os.remove(s2fname)
        if (s2fname != ""):
            try:
                wb = openpyxl.load_workbook(s2fname)  # 数据写入excel
            except:
                wb = openpyxl.Workbook()
                wb.save(s2fname)
            sheet = wb.active

            rows = sheet.max_row
            sheet.append(["序号", "指令集", "指令", "参数1","参数2","参数3","参数4","参数5"])
            A_1 = int(self.lineEdit_28.text())  # 表示要保存的工程参数的行数
            print(A_1)
            i = 1
            while i <= A_1:
                a = self.TableWidget.item(i - 1,0).text()
                b = self.TableWidget.item(i - 1,1).text()
                c = self.TableWidget.item(i - 1, 2).text()
                d = self.TableWidget.item(i - 1, 3).text()
                e = self.TableWidget.item(i - 1, 4).text()
                f = self.TableWidget.item(i - 1, 5).text()
                g = self.TableWidget.item(i - 1, 6).text()

                if a == "系统操作" and b == "停止":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 1, c, d, e, f, g])

                if a == "系统操作" and b == "启动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 2, c, d, e, f, g])

                if a == "系统操作" and b == "暂停":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 3, c, d, e, f, g])

                if a == "系统操作" and b == "恢复":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 4, c, d, e, f, g])

                if a == "系统操作" and b == "延时等待":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 5, c, d, e, f, g])

                if a == "系统操作" and b == "等待电机完成":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 6, c, d, e, f, g])

                if a == "系统操作" and b == "停止电机运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 7, c, d, e, f, g])

                if a == "系统操作" and b == "常等待":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,1, 8, c, d, e, f, g])

                if a == "流程控制" and b == "程序间跳转":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,2, 1, c, d, e, f, g])

                if a == "流程控制" and b == "程序循环":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,2, 2, c, d, e, f, g])

                if a == "流程控制" and b == "输入跳转":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,2, 3, c, d, e, f, g])

                if a == "流程控制" and b == "开启输入中断":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,2, 4, c, d, e, f, g])

                if a == "流程控制" and b == "关闭输入中断":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,2, 5, c, d, e, f, g])

                if a == "输出口设置" and b == "输出口设置":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,3, 1, c, d, e, f, g])

                if a == "回零运动" and b == "设置回零速度":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,4, 1, c, d, e, f, g])

                if a == "回零运动" and b == "启动回零":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,4, 2, c, d, e, f, g])

                if a == "插补运动" and b == "设置点位速度":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 1, c, d, e, f, g])

                if a == "插补运动" and b == "三轴相对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 2, c, d, e, f, g])

                if a == "插补运动" and b == "单轴绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 3, c, d, e, f, g])

                if a == "插补运动" and b == "XY绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 4, c, d, e, f, g])

                if a == "插补运动" and b == "XZ绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 5, c, d, e, f, g])

                if a == "插补运动" and b == "YZ绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 6, c, d, e, f, g])

                if a == "插补运动" and b == "三轴绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 7, c, d, e, f, g])

                if a == "插补运动" and b == "XY圆弧插补":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 8, c, d, e, f, g])

                if a == "插补运动" and b == "XZ圆弧插补":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 9, c, d, e, f, g])

                if a == "插补运动" and b == "YZ圆弧插补":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,5, 10, c, d, e, f, g])

                if a == "独立运动" and b == "独立运动速度":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,6, 1, c, d, e, f, g])

                if a == "独立运动" and b == "相对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,6, 2, c, d, e, f, g])

                if a == "独立运动" and b == "X绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,6, 3, c, d, e, f, g])

                if a == "独立运动" and b == "Y绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,6, 4, c, d, e, f, g])

                if a == "独立运动" and b == "Z绝对运动":
                    sheet.append([i,a, b, c, d, e, f, g])
                    sheet.append([i,6, 5, c, d, e, f, g])
                i += 1
            sheet.append([A_1])
            wb.save(s2fname)

    def openExcelfiles(self):
        excelname = QFileDialog.getOpenFileName(self,'打开工程',self.ProjectName ,"xlsx files (*.xlsx);;xls files (*.xls);;all files(*.*)")[0]
        # excelname = QFileDialog.getOpenFileName(self, '打开工程', self.ProjectName, "xlsx files (*.xlsx);;xls files (*.xls)")[0]
        print(excelname)
        if (excelname == ''):
            return

        data = xlrd.open_workbook(excelname)  # 打开目标文件
        print(data.sheet_names())  # 打印显示文件中各个sheet的名字
        table = data.sheets()[0]  # 获取第一个sheet的数据
        print(table)  #

#todo 如果要是修改的话 那A_1就是奇数
        A_1 = table.nrows  # 获取行数  （A_1肯定是奇数行）按照15行的设置
        print(A_1)

        B_1 = (A_1 - 1) / 2  #  B_1为 7
        print(B_1)
        self.lineEdit_28.setText(str(int(B_1)))  #将行数7写进文本框
        C_1 = B_1   #  7
        print(C_1)
        self.TableWidget.clearContents()  #清空表格的内容
        i = 1
        while i <= C_1:
            command_info_list_i = table.row_values(2 * i)  # 获取sheet中第i行的数据
            print(command_info_list_i)
            a=int(command_info_list_i[1])
            b=int(command_info_list_i[2])
            print(a)
            print(b)

            newItem_1 = QTableWidgetItem(str(command_info_list_i[3]))
            self.TableWidget.setItem(i-1,2, newItem_1)

            newItem_2 = QTableWidgetItem(str(command_info_list_i[4]))
            self.TableWidget.setItem(i-1,3,newItem_2)

            newItem_3 = QTableWidgetItem(str(command_info_list_i[5]))
            self.TableWidget.setItem(i-1, 4, newItem_3)

            newItem_4 = QTableWidgetItem(str(0))
            self.TableWidget.setItem(i-1, 5, newItem_4)

            newItem_5 = QTableWidgetItem(str(0))
            self.TableWidget.setItem(i-1, 6, newItem_5)

            comboBoxList1 = ["系统操作", "流程控制", "输出口设置", "回零运动", "插补运动", "独立运动"]
            comboBoxList2 = {"系统操作": ["停止", "启动", "暂停", "恢复", "延时等待", "等待电机完成", "停止电机运动", "常等待"],
                             "流程控制": ["程序间跳转", "程序循环", "输入跳转", "开启输入中断", "关闭输入中断"],
                             "输出口设置": ["输出口设置"],
                             "回零运动": ["设置回零速度", "启动回零"],
                             "插补运动": ["设置点位速度", "三轴相对运动", "单轴绝对运动", "XY绝对运动", "XZ绝对运动", "YZ绝对运动", "三轴绝对运动",
                                  "XY圆弧插补", "XZ圆弧插补", "YZ圆弧插补"],
                             "独立运动": ["独立运动速度", "相对运动", "X绝对运动", "Y绝对运动", "Z绝对运动"]}

            if a == 1:
                if b == 1:
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][0]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“停止”
                elif b == 2:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][1]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“启动”
                elif b == 3:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][2]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1, 1, newItem_7)

                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“暂停”
                elif b == 4:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][3]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“恢复”
                elif b == 5:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][4]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“延时等待”
                elif b == 6:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][5]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“等待电机完成”
                elif b == 7:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][6]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i , 1, newItem_7)

                    # 在指令集的下拉框中设置“系统操作”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“停止电机运动”
                elif b == 8:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["系统操作"][7]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)
            elif a == 2:
                if b == 1:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["流程控制"][0]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“流程控制”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“程序间跳转”
                elif b == 2:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["流程控制"][1]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“流程控制”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“程序循环”
                elif b == 3:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["流程控制"][2]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“流程控制”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“输入跳转”
                elif b == 4:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["流程控制"][3]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“流程控制”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“开启输入中断”
                elif b == 5:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["流程控制"][4]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)
                    # 在指令集的下拉框中设置“流程控制”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“关闭输入中断”
            elif a == 3:
                if b == 1:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["输出口设置"][0]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“输出口设置”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“输出口设置”
            elif a == 4:
                if b == 1:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1, 0, newItem_6)

                    ob = comboBoxList2["回零运动"][0]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“回零运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“设置回零速度”
                elif b == 2:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["回零运动"][1]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)
                    # 在指令集的下拉框中设置“回零运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“启动回零”
            elif a == 5:
                if b == 1:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][0]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)
                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“设置点位速度”
                elif b == 2:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][1]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“三轴相对运动”
                elif b == 3:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][2]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“单轴绝对运动”
                elif b == 4:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][3]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“XY绝对运动”
                elif b == 5:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1, 0, newItem_6)

                    ob = comboBoxList2["插补运动"][4]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“XZ绝对运动”
                elif b == 6:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][5]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1, 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“YZ绝对运动”
                elif b == 7:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][6]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1, 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“三轴绝对运动”
                elif b == 8:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1, 0, newItem_6)

                    ob = comboBoxList2["插补运动"][7]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“XZ绝对运动”
                elif b == 9:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][8]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1, 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“YZ绝对运动”
                elif b == 10:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["插补运动"][9]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1, 1, newItem_7)

                    # 在指令集的下拉框中设置“直线点位运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“三轴绝对运动”
            elif a == 6:
                if b == 1:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["独立运动"][0]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)

                    # 在指令集的下拉框中设置“独立插补运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“设置独立运动速度”
                elif b == 2:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1 , 0, newItem_6)

                    ob = comboBoxList2["独立运动"][1]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)
                    # 在指令集的下拉框中设置“独立插补运动”在设置的时候也要加一个循环，带有i就可以，
                    # 在指令的下拉框中设置“三轴独立插补”
                elif b == 3:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1, 0, newItem_6)

                    ob = comboBoxList2["独立运动"][2]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1 , 1, newItem_7)
                elif b == 4:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1, 0, newItem_6)

                    ob = comboBoxList2["独立运动"][3]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1, 1, newItem_7)
                elif b == 5:
                    print()
                    op = comboBoxList1[a - 1]
                    newItem_6 = QTableWidgetItem(str(op))
                    self.TableWidget.setItem(i-1, 0, newItem_6)

                    ob = comboBoxList2["独立运动"][4]
                    newItem_7 = QTableWidgetItem(str(ob))
                    self.TableWidget.setItem(i-1, 1, newItem_7)
            i += 1

    def Line_Arc_interpolation(self):
        self.verticalGroupBox_3.setVisible(False)
        self.GroupBoxPara.setVisible(False)
        self.GroupBoxEditProgram.setVisible(False)
        self.GroupBoxLine_arc.setVisible(True)
        if self.timer1>0:
            self.killTimer(self.timer1)
            self.timer1 = 0


if __name__ == '__main__':
    import cgitb
    cgitb.enable(format="text")  # 这两句可以避免PYQT5直接崩溃
    app = QtWidgets.QApplication(sys.argv)
    myshow = Pyqt5_Serial()
    myshow.show()
    myshow.activateWindow()
    sys.exit(app.exec_())

